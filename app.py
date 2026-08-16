import streamlit as st
import subprocess
import json
from pathlib import Path
from io import BytesIO
import re
import zipfile
from datetime import datetime
from streamlit_autorefresh import st_autorefresh
import openpyxl
import smtplib
from email.message import EmailMessage

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "bin"))

from dashboard_page import render_dashboard
from progress_tracker import render_progress_tracker
from qc_dashboard import render_qc_dashboard
from results_page import render_results_page

# ════════════════════════════════════════════════
# CONFIGURATION PAGE
# ════════════════════════════════════════════════
st.set_page_config(
    page_title="MaRS-py-upgrade · Pipeline FASTQ",
    layout="wide",
    page_icon="🧬"
)

REPO_DIR     = Path.home() / "MaRS-py-upgrade"
HISTORY_FILE = REPO_DIR / "runs_history.json"
HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)

# ════════════════════════════════════════════════
# TABLE DES MÉDICAMENTS — CODE TRAITEMENT
# ════════════════════════════════════════════════
DRUG_CODES = {
    "(AL) Arthemether-Lumafantrine - Coartem":              "A",
    "(CQ) Chloroquine":                                      "B",
    "(CQ+PQ) Chloroquine + Piperaquine":                    "C",
    "(AL+PQ) Arthemether-Lumafantrine + Piperaquine":       "D",
    "(AL+MQ) Arthemether + Mefloquine (Lariam)":            "E",
    "(AS+AQ) Artesunate + Amodiaquine":                     "F",
    "(AS+SP) Artesunate + Sulphadoxine-Pyrimethamine":      "G",
    "(AS+MQ) Artesunate + Mefloquine":                      "H",
    "(DHA+PQ) Dihydroartemisinin + Piperaquine":            "I",
    "(AS+SP+PQ) Artesunate + Sulphadoxine + Pyrimethamine": "J",
    "(AL+MQ+PQ) Arthemether + Mefloquine + Piperaquine":    "K",
    "Doxycycline":                                           "L",
    "(Malarone) Atovaquone/Proguanil":                      "M",
    "Quinine":                                               "N",
    "Quinine + Doxycydine":                                 "O",
    "Atovaquone/Proguanil + Arthemether-Lumafantrine":      "P",
    "Mefloquine":                                            "Q",
    "Vanomycin and Rocephin":                               "R",
    "Malarone + Doxycycline":                               "S",
    "Doxycycline, Malarone, Quinine and Artesunate":        "T",
    "Hydroxychloroquine":                                    "U",
    "Atovaquone":                                            "V",
    "Malarone + Coartem + Artesunate":                      "W",
    "Coartem + Artesunate":                                  "Y",
    "(AS+PY) Artesunate + Pyronaridine":                    "Z",
}
DRUG_OPTIONS = ["— Sélectionner un médicament —"] + list(DRUG_CODES.keys())


# ════════════════════════════════════════════════
# MARQUEURS MOLÉCULAIRES — LOGIQUE BIT CODE
# ════════════════════════════════════════════════
MOL_MARKERS_LIST = ["k13", "crt", "mdr1", "cytb", "dhps", "dhfr", "pfs47", "TBD1", "TBD2"]
MOL_MARKERS_BITS = {m: 8 - i for i, m in enumerate(MOL_MARKERS_LIST)}


def compute_mol_code(selected: list) -> str:
    decimal = sum(2 ** MOL_MARKERS_BITS[m] for m in selected if m in MOL_MARKERS_BITS)
    return str(decimal).zfill(3)


def mol_code_to_markers(code_str: str) -> list:
    try:
        decimal = int(code_str)
    except (ValueError, TypeError):
        return []
    selected = []
    for m in MOL_MARKERS_LIST:
        bit = MOL_MARKERS_BITS[m]
        if decimal & (1 << bit):
            selected.append(m)
    return selected


# ════════════════════════════════════════════════
# HELPERS — HISTORIQUE
# ════════════════════════════════════════════════

def load_history() -> list[dict]:
    if HISTORY_FILE.exists():
        try:
            history = json.loads(HISTORY_FILE.read_text())
        except Exception:
            return []
        
        updated = False
        for entry in history:
            if entry.get("status") == "running":
                run_id = entry["run_id"]
                # Vérifier si le ZIP existe → analyse terminée avec succès
                zip_path = Path(entry["zip_path"]) if entry.get("zip_path") else None
                paths = get_run_paths(run_id)
                if paths["zip"].exists():
                    entry["status"] = "success"
                    entry["zip_path"] = str(paths["zip"])
                    updated = True
                else:
                    # Vérifier si Docker tourne encore pour ce run
                    result = subprocess.run(
                        ["docker", "ps",
                        "--filter", f"label=mars_run_id={run_id}", "--format", "{{.ID}}"], capture_output=True, text=True, timeout=4,
                    )

                    docker_running = result.stdout.strip() != ""
                    if not docker_running:
                        # Docker arrêté sans ZIP → échec
                        entry["status"] = "failed"
                        updated = True
        
        if updated:
            HISTORY_FILE.write_text(json.dumps(history, indent=2, ensure_ascii=False))
        
        return history
    return []

def save_run_to_history(run_id, sample_names, status, duration_sec, zip_path):
    history = load_history()
    entry = {
        "run_id":       run_id,
        "date":         datetime.now().strftime("%Y-%m-%d %H:%M"),
        "samples":      len(sample_names),
        "sample_list":  [s["base"] for s in sample_names if isinstance(s, dict) and "base" in s],
        "status":       status,
        "duration_sec": duration_sec,
        "zip_path":     str(zip_path) if zip_path else None,
    }
    history = [h for h in history if h["run_id"] != run_id]
    history.insert(0, entry)
    history = history[:50]
    HISTORY_FILE.write_text(json.dumps(history, indent=2, ensure_ascii=False))


def delete_run_from_history(run_id):
    history = [h for h in load_history() if h["run_id"] != run_id]
    HISTORY_FILE.write_text(json.dumps(history, indent=2, ensure_ascii=False))


# ════════════════════════════════════════════════
# HELPERS — NOTIFICATIONS
# ════════════════════════════════════════════════
def send_email_notification(run_id, n_samples, duration_sec, success):
    try:
        email_cfg = st.secrets.get("EMAIL", {})
        if not email_cfg.get("enabled"):
            return
        from_addr = email_cfg.get("from")
        to_addr   = email_cfg.get("to")
        password  = email_cfg.get("password")
        if not (from_addr and to_addr and password):
            return
        dur_str = f"{duration_sec // 60} min {duration_sec % 60} s"
        status  = "Succès" if success else "Échec"
        msg = EmailMessage()
        msg["Subject"] = f"[MaRS-py-upgrade] Run {run_id} — {status}"
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg.set_content(
            f"Run {run_id} terminé — {status}\n"
            f"Échantillons analysés : {n_samples}\n"
            f"Durée : {dur_str}"
        )
        smtp_host = email_cfg.get("smtp_host", "smtp.gmail.com")
        smtp_port = email_cfg.get("smtp_port", 587)
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as smtp:
            smtp.starttls()
            smtp.login(from_addr, password)
            smtp.send_message(msg)
    except Exception:
        pass


def notify_pipeline_done(run_id, n_samples, duration_sec, success):
    if not st.session_state.get("notif_enabled", True):
        return
    if st.session_state.get("notif_failure_only") is True and success:
        return
    send_email_notification(run_id, n_samples, duration_sec, success)
    dur_str = f"{duration_sec // 60} min {duration_sec % 60} s"
    banner = {
        "type": "success" if success else "error",
        "msg": (
            f"✅ Run {run_id} terminé · {n_samples} échantillon(s) analysé(s) · ⏱ {dur_str}"
            if success else
            f"❌ Run **{run_id}** a échoué après {dur_str}"
        )
    }
    st.session_state["notif_banner"] = banner
    notif_file = REPO_DIR / "pending_notif.json"
    notif_file.write_text(json.dumps(banner, ensure_ascii=False))


# ════════════════════════════════════════════════
# HELPERS — QC PARSING
# ════════════════════════════════════════════════
def get_module_badge(status):
    colors = {
        "pass": ("✔", "#4ade80", "#0a1e14", "#22c55e33"),
        "warn": ("⚠", "#fbbf24", "#1a140a", "#f59e0b33"),
        "fail": ("✖", "#f87171", "#1a0a0e", "#ef444433"),
    }
    icon, color, bg, border = colors.get(status.lower(), ("?", "#94a3b8", "#0d1929", "#1e293b"))
    return (
        f'<span style="display:inline-flex;align-items:center;gap:4px;'
        f'font-size:0.72rem;font-weight:700;padding:3px 8px;border-radius:20px;'
        f'background:{bg};color:{color};border:1px solid {border};">'
        f'{icon} {status.upper()}</span>'
    )

@st.cache_data(ttl=300, show_spinner=False)
def get_qc_data_for_run(run_id):
    paths = get_run_paths(run_id)
    qc_data = {}
    out_dir = paths["output"]
    if not out_dir.exists():
        return qc_data
    for html_file in out_dir.glob("**/*_fastqc.html"):
        sample_name = html_file.stem.replace("_fastqc", "")
        if sample_name not in qc_data:
            qc_data[sample_name] = {"html_files": [], "txt_metrics": {}}
        qc_data[sample_name]["html_files"].append(html_file)
    for zip_file in out_dir.glob("**/*_fastqc.zip"):
        sample_name = zip_file.stem.replace("_fastqc", "")
        if sample_name not in qc_data:
            qc_data[sample_name] = {"html_files": [], "txt_metrics": {}}
        try:
            with zipfile.ZipFile(zip_file, "r") as zf:
                txt_candidates = [n for n in zf.namelist() if n.endswith("fastqc_data.txt")]
                if txt_candidates:
                    with zf.open(txt_candidates[0]) as f:
                        content = f.read().decode("utf-8", errors="replace")
                    metrics = {"total_reads": "—", "read_length": "—", "gc_pct": "—", "modules": {}}
                    for line in content.splitlines():
                        if line.startswith(">>") and not line.startswith(">>END"):
                            parts = line.lstrip(">").split("\t")
                            if len(parts) == 2:
                                metrics["modules"][parts[0].strip()] = parts[1].strip()
                        elif line.startswith("Total Sequences"):
                            n = int(line.split("\t")[1].strip())
                            metrics["total_reads"] = f"{n:,}"
                        elif line.startswith("Sequence length"):
                            metrics["read_length"] = line.split("\t")[1].strip() + " bp"
                        elif line.startswith("%GC"):
                            metrics["gc_pct"] = line.split("\t")[1].strip() + "%"
                    if not qc_data[sample_name]["txt_metrics"] or "_R1" in zip_file.name:
                        qc_data[sample_name]["txt_metrics"] = metrics
        except Exception:
            pass
    return qc_data


# ════════════════════════════════════════════════
# HELPERS — CHEMINS & VALIDATION
# ════════════════════════════════════════════════
def get_run_paths(run_id):
    base_dir = REPO_DIR / "runs" / f"run_{run_id}"
    return {
        "base":   base_dir,
        "input":  base_dir / "data",
        "output": base_dir / "output",
        "logs":   base_dir / "logs",
        "zip":    base_dir / "resultats_pipeline.zip",
    }


def safe_filename(name):
    return Path(name).name


def only_digits(value, length):
    return value.isdigit() and len(value) == length


def only_letters(value, length):
    return value.isalpha() and len(value) == length


def only_upper_letter(value, length=1):
    return value.isalpha() and len(value) == length


def validate_common(year, country, state, day, treat, mol, proc):
    errors = []
    if not only_digits(year, 2):        errors.append("Année : exactement 2 chiffres (ex: 24)")
    if not only_letters(country, 2):    errors.append("Pays : exactement 2 lettres (ex: SN)")
    if not only_letters(state, 2):      errors.append("État/Province : exactement 2 lettres (ex: DK)")
    if not only_digits(day, 2):         errors.append("Jour traitement : exactement 2 chiffres (ex: 00)")
    if not only_upper_letter(treat, 1): errors.append("Code traitement : exactement 1 lettre (ex: A)")
    if not only_digits(mol, 3):         errors.append("Mol markers : exactement 3 chiffres (ex: 000)")
    if proc not in ("1", "2"):          errors.append("Processé : 1 (1er passage) ou 2 (répétition)")
    return errors


def build_base(year, country, state, day, treat, samp_part, seq_part):
    common = (year + country + state + day + treat).upper()
    return common + samp_part + seq_part


def get_suffix_from_original(original_name):
    stem = original_name
    if stem.endswith(".fastq.gz"):
        stem = stem[:-9]
        ext = ".fastq.gz"
    elif stem.endswith(".fastq"):
        stem = stem[:-6]
        ext = ".fastq"
    else:
        ext = ""
    idx = stem.find("_")
    if idx == -1:
        return ext
    return stem[idx:] + ext


def build_new_filename(amd_base, original_name):
    suffix = get_suffix_from_original(original_name)
    return amd_base + suffix


def get_sample_id(filename: str) -> str:
    """Extract the part before the first underscore as the Sample ID."""
    return filename.split("_")[0] if "_" in filename else filename.split(".")[0]

def get_pair_prefix(filename: str) -> str:
    """
    Extrait le préfixe commun R1/R2 en retirant _R1/_R2 (et variantes)
    ainsi que l'extension .fastq.gz ou .fastq.
    Ex: SAMPLE001_S1_L001_R1_001.fastq.gz → SAMPLE001_S1_L001__001
    """
    name = filename
    # Retirer extension
    if name.endswith(".fastq.gz"):
        name = name[:-9]
    elif name.endswith(".fastq"):
        name = name[:-6]
    # Neutraliser _R1_ / _R2_ / .R1. / .R2. / _1 / _2 en fin
    name = re.sub(r'[_.]R[12][_.]', '__', name, flags=re.IGNORECASE)
    name = re.sub(r'[_.]R[12]$',    '',   name, flags=re.IGNORECASE)
    name = re.sub(r'_[12]$',        '',   name)
    return name

def reset_state():
    preserve = {"authenticated", "notif_enabled", "notif_failure_only"}
    for k, v in DEFAULTS.items():
        if k not in preserve:
            st.session_state[k] = v() if callable(v) else v


# ════════════════════════════════════════════════
# EXCEL TEMPLATE GENERATOR
# ════════════════════════════════════════════════
def generate_template(pairs=None):
    """
    Génère le template Excel.
    Si `pairs` est fourni (liste de dicts {"r1": file, "r2": file}),
    la colonne ID Sample est pré-remplie avec les vrais IDs des fichiers.
    Sinon, des exemples génériques sont utilisés.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nomenclature AMD"
 
    headers = [
        "ID Sample",
        "Type",
        "Année (2 chiffres)",
        "Pays (2 lettres)",
        "État/Province (2 lettres)",
        "Jour traitement (2 chiffres)",
        "Code traitement (1 lettre)",
        "Sample ID Individual (4 chiffres)",
        "Sample ID Pooled (3 chiffres)",
        "Nb de pool (2 chiffres)",
        "k13 (0/1)", "crt (0/1)", "mdr1 (0/1)",
        "cytb (0/1)", "dhps (0/1)", "dhfr (0/1)",
        "pfs47 (0/1)", "TBD1 (0/1)", "TBD2 (0/1)",
        "Processé (1=1er, 2=répét.)",
    ]
    ws.append(headers)
 
    from openpyxl.styles import Font, PatternFill, Alignment
    id_fill      = PatternFill(start_color="0f4c81", end_color="0f4c81", fill_type="solid")
    header_fill  = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
    mol_hdr_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
 
    for col_num, cell in enumerate(ws[1], 1):
        is_id  = col_num == 1
        is_mol = 11 <= col_num <= 19
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = id_fill if is_id else (mol_hdr_fill if is_mol else header_fill)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = (
            20 if is_id else (18 if is_mol else 26)
        )
 
    # ── Remplissage des lignes de données ──────────────────────────────────
    # Si des paires réelles sont fournies, on pré-remplit ID Sample + lignes vides
    # Sinon, on met des exemples génériques
    if pairs:
        data_fill = PatternFill(start_color="f0f9ff", end_color="f0f9ff", fill_type="solid")
        alt_fill  = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
        for idx, pair in enumerate(pairs):
            sample_id = get_sample_id(pair["r1"].name)
            # Ligne avec ID Sample pré-rempli, reste vide à compléter
            row = [sample_id] + [""] * 19
            ws.append(row)
            fill = data_fill if idx % 2 == 0 else alt_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill
            # Mettre en gras + bleu l'ID Sample
            ws[ws.max_row][0].font = Font(bold=True, color="0f4c81")
    else:
        # Exemples génériques si pas de paires
        ex_fill = PatternFill(start_color="f0f9ff", end_color="f0f9ff", fill_type="solid")
        examples = [
            ["SAMPLE001", "Individual", "24", "SN", "DK", "00", "A", "1000", "", "",    1, 1, 1, 1, 1, 1, 1, 0, 0, "1"],
            ["SAMPLE002", "Individual", "24", "SN", "DK", "00", "B", "1001", "", "",    1, 0, 0, 0, 0, 0, 0, 0, 0, "1"],
            ["POOL001",   "Pooled",     "24", "SN", "DK", "00", "F", "",     "000","05",0, 0, 0, 0, 1, 1, 0, 0, 0, "1"],
        ]
        for row in examples:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.fill = ex_fill
 
    # ── Feuille Instructions ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Champ", "Valeur attendue", "Exemple"],
        ["ID Sample", "Identifiant original de l'échantillon (partie avant le '_' du nom de fichier — NE PAS MODIFIER)", "SAMPLE001"],
        ["Type", "Individual ou Pooled", "Individual"],
        ["Année", "2 chiffres", "24"],
        ["Pays", "2 lettres majuscules", "SN"],
        ["État/Province", "2 lettres majuscules", "DK"],
        ["Jour traitement", "2 chiffres", "00"],
        ["Code traitement", "1 lettre (voir feuille Médicaments)", "A"],
        ["Sample ID Individual", "4 chiffres (si Individual)", "1000"],
        ["Sample ID numérique Pooled", "3 chiffres (si Pooled)", "000"],
        ["Nb dans pool", "2 chiffres (si Pooled)", "05"],
        ["k13 … TBD2", "0 (absent) ou 1 (présent) pour chaque marqueur", "1"],
        ["", "", ""],
        ["NOTE", "⚠️ Ne pas modifier la colonne ID Sample — elle sert à identifier vos fichiers.", ""],
        ["NOTE", "Le code Mol markers est calculé automatiquement depuis les colonnes 0/1.", ""],
        ["NOTE", "Bit code : k13=bit8, crt=bit7, mdr1=bit6, cytb=bit5, dhps=bit4, dhfr=bit3, pfs47=bit2, TBD1=bit1, TBD2=bit0", ""],
        ["NOTE", "Exemple k13+crt+mdr1+cytb+dhps+dhfr+pfs47 → 111111100 → 508", ""],
        ["Processé", "1 (1er passage) ou 2 (répétition)", "1"],
    ]
    for row in instructions:
        ws2.append(row)
    for col_num, cell in enumerate(ws2[1], 1):
        cell.font = Font(bold=True)
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 60
 
    # ── Feuille mol markers ────────────────────────────────────────────────
    ws_mol = wb.create_sheet("mol markers")
    ws_mol.append(["Marqueur", "Position bit", "Valeur décimale"])
    mol_hdr = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
    for cell in ws_mol[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = mol_hdr
        cell.alignment = Alignment(horizontal="center")
    ws_mol.column_dimensions["A"].width = 16
    ws_mol.column_dimensions["B"].width = 16
    ws_mol.column_dimensions["C"].width = 20
    alt_fill_mol = PatternFill(start_color="faf5ff", end_color="faf5ff", fill_type="solid")
    for idx, m in enumerate(MOL_MARKERS_LIST):
        bit = MOL_MARKERS_BITS[m]
        ws_mol.append([m, f"bit {bit}", 2**bit])
        if idx % 2 == 0:
            for cell in ws_mol[ws_mol.max_row]:
                cell.fill = alt_fill_mol
 
    # ── Feuille Médicaments ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Médicaments")
    ws3.append(["Médicament", "Code traitement"])
    ws3_hfill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
    for cell in ws3[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = ws3_hfill
        cell.alignment = Alignment(horizontal="center")
    ws3.column_dimensions["A"].width = 55
    ws3.column_dimensions["B"].width = 20
    alt_fill2 = PatternFill(start_color="faf5ff", end_color="faf5ff", fill_type="solid")
    for idx, (drug, code) in enumerate(DRUG_CODES.items()):
        ws3.append([drug, code])
        if idx % 2 == 0:
            for cell in ws3[ws3.max_row]:
                cell.fill = alt_fill2
 
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════
# SESSION STATE
# ════════════════════════════════════════════════
DEFAULTS = {
    "authenticated":      False,
    "active_page":        "home",
    "prev_page":          "home",
    "uploaded_pairs":     list,
    "upload_done":        False,
    "naming_mode":        None,
    "sample_names":       list,
    "nomenclature_done":  False,
    "run_id":             None,
    "running":            False,
    "pipeline_done":      False,
    "zip_created":        False,
    "show_download":      False,
    "docker_started":     False,
    "launch_time":        0,
    "log_file":           "",
    "trigger_run":        False,
    "names_for_history":  list,
    "notif_enabled":      True,
    "notif_failure_only": False,
    "confirm_delete":     None,
    "qc_run_id":          None,
    "results_run_id":     None,
    "_pdf_bytes":         None,
    "_pdf_run_id":        None,
    "notif_banner":       None,
    "editing_sample":     None,
}

for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v() if callable(v) else v

_notif_file = REPO_DIR / "pending_notif.json"
if not st.session_state.get("notif_banner") and _notif_file.exists():
    try:
        st.session_state["notif_banner"] = json.loads(_notif_file.read_text())
    except Exception:
        pass


# ════════════════════════════════════════════════
# STYLE GLOBAL
# ════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@300;400;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, .stApp {
    background: #f5f5f5 !important;
    font-family: 'Source Sans Pro', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    color: #333333 !important;
    font-size: 15px;
}
.block-container { padding: 1.2rem 2rem !important; max-width: 1200px; }
hr { border: none !important; border-top: 1px solid #d8d8d8 !important; margin: 16px 0 !important; }

[data-testid="stSidebar"] {
    background: #1a2a4a !important;
    border-right: 1px solid #162240 !important;
}
[data-testid="stSidebar"] * { color: #ccd6f6 !important; font-family: 'Source Sans Pro', sans-serif !important; }
[data-testid="stSidebar"] h2 { color: #ffffff !important; font-size: 1rem !important; font-weight: 700 !important; }
[data-testid="stSidebar"] hr { border-top: 1px solid #2e3f6e !important; }
[data-testid="stSidebar"] .stButton > button {
    background: #2c4a7c !important; color: #e8f0fe !important;
    border: 1px solid #3a5a94 !important; border-radius: 4px !important; text-align: center !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #2c6fad !important; border-color: #2c6fad !important; color: #ffffff !important;
}

h1, h2, h3 { font-family: 'Source Sans Pro', sans-serif !important; color: #333333 !important; font-weight: 600 !important; }
h3 { font-size: 0.75rem !important; font-weight: 700 !important; letter-spacing: 1.5px !important; text-transform: uppercase !important; color: #2c6fad !important; margin-bottom: 10px !important; }

.mars-banner {
    position: relative; overflow: hidden;
    background: linear-gradient(135deg, #1a2a4a 0%, #2c6fad 100%);
    border: none; border-radius: 6px; padding: 28px 40px; margin-bottom: 20px;
    display: flex; align-items: center; justify-content: space-between; gap: 20px;
    box-shadow: 0 2px 8px rgba(26,42,74,0.3);
}
.mars-banner::before {
    content: ''; position: absolute; top: -50px; right: -50px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(255,255,255,0.06) 0%, transparent 70%);
    pointer-events: none;
}
.banner-title { font-family: 'Source Sans Pro', sans-serif; font-size: 2rem; color: #ffffff; font-weight: 700; letter-spacing: -0.3px; line-height: 1.2; margin: 0 0 6px; }
.banner-title span { color: #a8d4ff; }
.banner-sub { color: rgba(255,255,255,0.72); font-size: 0.87rem; line-height: 1.55; margin: 0; }
.banner-dna { font-size: 3.5rem; opacity: 0.35; flex-shrink: 0; }

.section-label {
    font-size: 0.75rem; font-weight: 700; letter-spacing: 2px;
    text-transform: uppercase; color: #888888;
    margin-bottom: 10px; display: flex; align-items: center; gap: 8px;
}
.section-label::after { content: ''; flex: 1; height: 1px; background: #d8d8d8; }

.wf-step {
    flex: 1; display: flex; align-items: center; gap: 10px;
    background: #ffffff; border: 1px solid #d8d8d8; border-right: none;
    padding: 12px 16px; box-shadow: 0 1px 2px rgba(0,0,0,0.05);
}
.wf-step:first-child { border-radius: 4px 0 0 4px; }
.wf-step:last-child  { border-right: 1px solid #d8d8d8; border-radius: 0 4px 4px 0; }
.wf-step.active { background: #e8f2fb; border-color: #2c6fad; }
.wf-step.done   { background: #f0fdf4; border-color: #7ec8a0; }
.wf-num {
    width: 28px; height: 28px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 0.75rem; font-weight: 700;
    background: #e0e0e0; color: #666666; flex-shrink: 0;
}
.wf-step.active .wf-num { background: #2c6fad; color: #ffffff; }
.wf-step.done   .wf-num { background: #28a745; color: #ffffff; }
.wf-label { font-size: 0.87rem; color: #666666; font-weight: 500; }
.wf-step.active .wf-label { color: #2c6fad; font-weight: 600; }
.wf-step.done   .wf-label { color: #1a7a3c; }

.stButton > button {
    background: #ffffff !important; color: #333333 !important;
    border: 1px solid #aaaaaa !important; border-radius: 3px !important;
    padding: 0.55rem 0.9rem !important; font-family: 'Source Sans Pro', sans-serif !important;
    font-weight: 400 !important; font-size: 0.92rem !important; text-align: left !important;
    transition: all 0.1s ease !important; line-height: 1.4 !important;
    box-shadow: 0 1px 1px rgba(0,0,0,0.06) !important;
}
.stButton > button:hover {
    background: #e8f2fb !important; border-color: #2c6fad !important;
    color: #2c4a7c !important; box-shadow: 0 1px 3px rgba(44,111,173,0.2) !important;
}

.stTextInput > label, .stRadio > label, .stFileUploader > label {
    color: #555555 !important; font-size: 0.88rem !important;
    font-weight: 600 !important; letter-spacing: 0.2px !important;
}
.stTextInput > div > div > input {
    background: #ffffff !important; border: 1px solid #aaaaaa !important;
    color: #333333 !important; border-radius: 3px !important;
    font-family: 'JetBrains Mono', monospace !important; font-size: 0.9rem !important;
    padding: 8px 10px !important; transition: border-color 0.1s, box-shadow 0.1s !important;
    box-shadow: inset 0 1px 2px rgba(0,0,0,0.06) !important;
}
.stTextInput > div > div > input:focus {
    border-color: #2c6fad !important; box-shadow: 0 0 0 2px rgba(44,111,173,0.2) !important; outline: none !important;
}
.stTextInput > div > div > input:disabled {
    background: #f2f2f2 !important; color: #aaaaaa !important; border-color: #dddddd !important;
}

[data-testid="stFileUploader"] {
    border: 2px dashed #aaaaaa !important; border-radius: 4px !important;
    padding: 18px !important; background: #fafafa !important; transition: border-color 0.1s !important;
}
[data-testid="stFileUploader"]:hover { border-color: #2c6fad !important; }

[data-testid="stExpander"] {
    background: #ffffff !important; border: 1px solid #d8d8d8 !important;
    border-radius: 4px !important; box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
}
[data-testid="stExpander"] summary { color: #333333 !important; font-weight: 600 !important; font-size: 0.88rem !important; }
[data-testid="stForm"] { background: #fafafa; border: 1px solid #d8d8d8; border-radius: 4px; padding: 16px !important; }

.stInfo, [data-testid="stNotification"] {
    background: #e8f2fb !important; border: 1px solid #bdd5ef !important;
    border-left: 4px solid #2c6fad !important; border-radius: 4px !important;
    color: #1a3a6a !important; font-size: 0.88rem !important;
}
.stSuccess {
    background: #edfaf0 !important; border: 1px solid #a8d5b5 !important;
    border-left: 4px solid #28a745 !important; border-radius: 4px !important;
    color: #155724 !important; font-size: 0.88rem !important;
}
.stError {
    background: #fdf0f0 !important; border: 1px solid #f0b8b8 !important;
    border-left: 4px solid #dc3545 !important; border-radius: 4px !important;
    color: #721c24 !important; font-size: 0.88rem !important;
}
.stWarning {
    background: #fff8e6 !important; border: 1px solid #ffd980 !important;
    border-left: 4px solid #ffc107 !important; border-radius: 4px !important;
    color: #664d03 !important; font-size: 0.88rem !important;
}

.stCodeBlock, pre, code {
    background: #f2f2f2 !important; border: 1px solid #d8d8d8 !important;
    border-radius: 3px !important; font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.8rem !important; color: #333333 !important;
}

/* ── CHANGE 1: Upload summary table ── */
.upload-table {
    width: 100%; border-collapse: collapse; margin: 14px 0;
    font-size: 0.85rem; background: #ffffff;
    border: 1px solid #d8d8d8; border-radius: 4px; overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.upload-table th {
    background: #1a2a4a; color: #ffffff; font-weight: 700;
    font-size: 0.72rem; letter-spacing: 1.5px; text-transform: uppercase;
    padding: 10px 14px; text-align: left;
}
.upload-table th.col-id   { background: #0f4c81; min-width: 110px; }
.upload-table th.col-fwd  { background: #1a5276; }
.upload-table th.col-rev  { background: #4a235a; }
.upload-table td {
    padding: 9px 14px; border-bottom: 1px solid #eeeeee;
    font-family: 'JetBrains Mono', monospace; font-size: 0.8rem;
    vertical-align: middle;
}
.upload-table tr:last-child td { border-bottom: none; }
.upload-table tr:hover td { background: #f0f6ff; }
.upload-table td.col-id   { font-weight: 700; color: #1a2a4a; }
.upload-table td.col-fwd  { color: #2c6fad; }
.upload-table td.col-rev  { color: #6f42c1; }
.upload-table td.col-idx  { color: #aaaaaa; font-size: 0.72rem; width: 32px; text-align: center; }

/* ── CHANGE 2: Manual nomenclature table ── */
.nom-table-wrap { overflow-x: auto; margin: 14px 0; }
.nom-table {
    width: 100%; border-collapse: collapse;
    font-size: 0.83rem; background: #ffffff;
    border: 1px solid #d8d8d8; border-radius: 4px; overflow: hidden;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.nom-table th {
    background: #1a2a4a; color: #ffffff; font-weight: 700;
    font-size: 0.68rem; letter-spacing: 1.2px; text-transform: uppercase;
    padding: 9px 10px; text-align: center; white-space: nowrap;
}
.nom-table th.col-id   { background: #0f4c81; }
.nom-table th.col-mol  { background: #4a235a; }
.nom-table td {
    padding: 6px 8px; border-bottom: 1px solid #eeeeee;
    vertical-align: middle; text-align: center;
}
.nom-table tr:last-child td { border-bottom: none; }
.nom-table tr:hover td { background: #f8f8ff; }
.nom-table td.id-cell  {
    font-family: 'JetBrains Mono', monospace; font-size: 0.82rem;
    font-weight: 700; color: #1a2a4a; text-align: left; min-width: 100px;
}
.apply-all-btn {
    font-size: 0.68rem !important; padding: 2px 6px !important;
    color: #2c6fad !important; background: #e8f2fb !important;
    border: 1px solid #bdd5ef !important; border-radius: 3px !important;
    cursor: pointer; white-space: nowrap;
}

.pair-card {
    background: #ffffff; border: 1px solid #d8d8d8; border-radius: 4px;
    padding: 10px 14px; margin-bottom: 6px; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    transition: border-color 0.1s;
}
.pair-card:hover { border-color: #2c6fad; }
.pair-card-header { display: flex; align-items: center; gap: 8px; margin-bottom: 4px; }
.pair-idx {
    font-size: 0.72rem; font-weight: 700; color: #ffffff;
    background: #2c6fad; padding: 2px 8px; border-radius: 3px;
    text-transform: uppercase; letter-spacing: 0.5px;
}
.pair-file {
    font-family: 'JetBrains Mono', monospace; font-size: 0.82rem;
    color: #333333; line-height: 1.5; display: flex; align-items: center; gap: 6px; flex-wrap: wrap;
}
.pair-file .r1 { color: #2c6fad; font-weight: 500; }
.pair-file .r2 { color: #6f42c1; font-weight: 500; }
.pair-file .sep { color: #999999; font-weight: 700; margin: 0 4px; }
.pair-file .sz  { color: #aaaaaa; font-size: 0.72rem; }

.assoc-row {
    background: #ffffff; border: 1px solid #d8d8d8; border-radius: 4px;
    padding: 10px 14px; margin-bottom: 5px; box-shadow: 0 1px 1px rgba(0,0,0,0.03);
}
.assoc-orig { font-family: 'JetBrains Mono', monospace; font-size: 0.83rem; color: #888888; }
.assoc-name { font-family: 'JetBrains Mono', monospace; font-size: 0.88rem; color: #2c6fad; font-weight: 600; }

.sample-card {
    background: #ffffff; border: 1px solid #d8d8d8; border-radius: 4px;
    padding: 10px 14px; margin-bottom: 6px; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.sample-card-title { font-weight: 600; color: #333333; font-size: 0.88rem; margin-bottom: 3px; }
.sample-card-file  { font-family: 'JetBrains Mono', monospace; font-size: 0.77rem; color: #666666; line-height: 1.55; }

.progress-wrap { background: #e0e0e0; border-radius: 3px; padding: 2px; margin: 8px 0 12px; border: 1px solid #cccccc; }
.progress-fill {
    border-radius: 2px; padding: 7px 14px; text-align: center; color: #ffffff;
    font-weight: 600; font-size: 0.8rem; letter-spacing: 0.3px;
    font-family: 'Source Sans Pro', sans-serif;
    transition: width 0.5s ease; background: #2c6fad; min-width: 50px;
}

.stDownloadButton > button {
    background: #2c6fad !important; color: #ffffff !important;
    border: 1px solid #245d94 !important; border-radius: 3px !important;
    padding: 0.6rem 1rem !important; font-weight: 600 !important;
    font-size: 0.9rem !important; text-align: center !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.1) !important; transition: background 0.1s !important;
}
.stDownloadButton > button:hover { background: #245d94 !important; border-color: #1e4f80 !important; }

.main-nav {
    display: flex; gap: 2px; margin-bottom: 1.2rem;
    background: #ffffff; border: 1px solid #d8d8d8;
    border-radius: 4px; padding: 4px; box-shadow: 0 1px 2px rgba(0,0,0,0.05);
}

.hist-card {
    background: #ffffff; border: 1px solid #d8d8d8; border-radius: 4px;
    padding: 12px 16px; margin-bottom: 6px;
    display: flex; align-items: center; gap: 14px;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04); transition: border-color 0.1s;
}
.hist-card:hover { border-color: #2c6fad; }
.hist-badge { font-size: 0.72rem; font-weight: 700; letter-spacing: 0.3px; padding: 2px 9px; border-radius: 3px; }
.badge-success { background: #edfaf0; color: #155724; border: 1px solid #a8d5b5; }
.badge-failed  { background: #fdf0f0; color: #721c24; border: 1px solid #f0b8b8; }
.badge-running { background: #e8f2fb; color: #1a3a6a; border: 1px solid #bdd5ef; }

.qc-metric {
    background: #ffffff; border: 1px solid #d8d8d8; border-radius: 4px;
    padding: 12px 16px; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.qc-label { font-size: 0.75rem; font-weight: 700; letter-spacing: 1.5px; text-transform: uppercase; color: #888888; margin-bottom: 5px; }
.qc-value { font-size: 1.45rem; font-weight: 700; color: #1a2a4a; font-family: 'JetBrains Mono', monospace; }

.mol-selector-wrap {
    background: #fafafa; border: 1px solid #d8d8d8; border-radius: 4px;
    padding: 12px 14px; margin-bottom: 10px;
}
.mol-bit-display { display: flex; gap: 4px; flex-wrap: wrap; margin: 8px 0 4px; }
.mol-bit { display: inline-flex; flex-direction: column; align-items: center; gap: 2px; min-width: 36px; }
.mol-bit-box {
    width: 34px; height: 26px; border-radius: 4px;
    display: flex; align-items: center; justify-content: center;
    font-size: 13px; font-weight: 700; font-family: 'JetBrains Mono', monospace;
    border: 1px solid #d8d8d8; background: #ffffff; color: #cccccc;
}
.mol-bit-box.on { background: #e8f2fb; color: #2c6fad; border-color: #2c6fad; }
.mol-bit-name { font-size: 10px; color: #888888; font-family: 'JetBrains Mono', monospace; }
.mol-decimal { font-size: 1.3rem; font-weight: 700; color: #2c6fad; font-family: 'JetBrains Mono', monospace; margin-left: 8px; }

input[type=number]::-webkit-inner-spin-button,
input[type=number]::-webkit-outer-spin-button { -webkit-appearance: none; margin: 0; }

/* ── Masquer toolbar Streamlit native ── */
[data-testid="stToolbar"] { display: none !important; }
header[data-testid="stHeader"] { display: none !important; }

/* ── Header fixe ── */
.mars-fixed-header {
    position: fixed;
    top: 0; left: 244px; right: 0;
    z-index: 9999;
    background: #f0f2f6;
    padding: 12px 2rem 8px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}
.mars-fixed-header .mars-banner {
    margin-bottom: 8px;
    padding: 20px 40px;
}

/* ── Compensation contenu sous le header ── */
section[data-testid="stMain"] > div:first-child {
    padding-top: 220px !important;
}
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════
# WIDGET AFFICHAGE BIT CODE MOL MARKERS
# ════════════════════════════════════════════════
def render_mol_bitcode(selected_markers: list, mol_code: str):
    bits_html = ""
    for m in MOL_MARKERS_LIST:
        on = "on" if m in selected_markers else ""
        val = "1" if m in selected_markers else "0"
        bits_html += f'<div class="mol-bit"><div class="mol-bit-box {on}">{val}</div><div class="mol-bit-name">{m}</div></div>'
    decimal_val = int(mol_code) if mol_code.isdigit() else 0
    st.markdown(f"""
    <div class="mol-selector-wrap">
        <div style="font-size:0.72rem;font-weight:700;color:#888888;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:6px;">
            Bit code → décimal
        </div>
        <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;">
            <div class="mol-bit-display">{bits_html}</div>
            <span style="font-size:1.2rem;color:#aaaaaa;font-weight:700;margin:0 4px;">→</span>
            <div>
                <span style="font-size:0.72rem;color:#888888;display:block;margin-bottom:2px;">Mol markers</span>
                <span class="mol-decimal">{mol_code}</span>
                <span style="font-size:0.75rem;color:#888888;margin-left:4px;">(={decimal_val})</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════
with st.sidebar:
    st.markdown('<div style="display:flex;justify-content:center;padding:12px 0 4px;">', unsafe_allow_html=True)
    try:
        st.image("images/bioinf.jpeg", width="stretch")
    except Exception:
        st.markdown("🧬")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("## MaRS-py-upgrade")
    st.markdown("""
    <div style="font-size:0.85rem;color:#ccd6f6;line-height:1.7;">
    Pipeline bioinformatique pour :<br>
    🔹 Analyse FASTQ<br>🔹 Alignement<br>🔹 Variant Calling<br>
    🔹 Annotation<br>🔹 Haplotype<br>🔹 Rapport final
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("<br>", unsafe_allow_html=True)

    if st.session_state.get("authenticated", False):
        if st.button("🚪 Déconnexion", width='stretch'):
            if st.session_state.get("running", False):
                id_result = subprocess.run(["docker", "ps", "--filter", "ancestor=bioinfo_pipeline", "--format", "{{.ID}}"], capture_output=True, text=True)
                container_id = id_result.stdout.strip()
                if container_id:
                    subprocess.run(["docker", "stop", container_id], capture_output=True)
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.session_state["authenticated"] = False
            st.rerun()

if st.session_state.get("running"):
    st_autorefresh(interval=5000, key="pipeline_refresh")


# ════════════════════════════════════════════════
# AUTHENTIFICATION
# ════════════════════════════════════════════════
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    st.markdown("""
    <style>
    html,body,.stApp{background:#ffffff !important;}
    .login-wrap{max-width:420px;margin:80px auto 0;text-align:center;}
    .login-logo{font-family:'DM Serif Display',serif;font-size:2.5rem;font-weight:700;white-space:nowrap;background:linear-gradient(135deg,#38bdf8,#818cf8);-webkit-background-clip:text;-webkit-text-fill-color:transparent;letter-spacing:-1px;margin-bottom:0.25rem;}
    .login-sub{color:#6c757d;font-family:'DM Sans',sans-serif;font-size:0.9rem;margin-bottom:2.5rem;}
    .stTextInput input{background:#ffffff !important;border:1px solid #aaaaaa !important;color:#333333 !important;border-radius:4px !important;padding:14px 16px !important;}
    .stTextInput input:focus{border-color:#38bdf8 !important;box-shadow:0 0 0 3px rgba(56,189,248,0.12) !important;}
    .stButton>button{background:linear-gradient(135deg,#1f70b8,#7c3aed) !important;color:white !important;border:none !important;border-radius:10px !important;font-weight:600 !important;padding:0.75rem !important;}
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div class="login-wrap">
        <div class="login-logo">🧬 MaRS-py-upgrade</div>
        <div class="login-sub">Pipeline bioinformatique FASTQ → Rapport</div>
    </div>
    """, unsafe_allow_html=True)
    col = st.columns([1, 2, 1])[1]
    with col:
        pwd = st.text_input("Mot de passe", type="password", label_visibility="collapsed", placeholder="🔒  Mot de passe")
        if st.button("Connexion", width='stretch'):
            if pwd == st.secrets["APP_PASSWORD"]:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Mot de passe incorrect")
    st.stop()


# ════════════════════════════════════════════════
# HEADER FIXE (bannière + navigation)
# ════════════════════════════════════════════════
st.markdown(f"""
<div class="mars-fixed-header">
    <div class="mars-banner">
        <div>
            <div class="banner-title">🧬 <span>MaRS-py</span>-upgrade</div>
            <p class="banner-sub">
                Pipeline bioinformatique pour l'analyse de séquences FASTQ
                de <em>Plasmodium falciparum</em><br>
                ● FASTQ → Alignement → Variant Calling → Annotation → Haplotype → Rapport
            </p>
        </div>
        <div class="banner-dna">🧬</div>
    </div>
</div>
""", unsafe_allow_html=True)

# Navigation avec de vrais boutons Streamlit (pas de liens HTML)
_active_page = st.session_state.get("active_page", "home")
_nav_items = [
    ("home",     "🏠 Accueil"),
    ("pipeline", "🔬 Nouveau pipeline"),
    ("results",  "🧬 Résultats"),
    ("history",  "📋 Historique"),
    ("settings", "⚙️ Paramètres"),
]

nav_cols = st.columns(len(_nav_items))
for col, (_pid, _lbl) in zip(nav_cols, _nav_items):
    with col:
        if st.button(_lbl, key=f"nav_{_pid}", width='stretch'):
            st.session_state["prev_page"]   = st.session_state.get("active_page", "home")
            st.session_state["active_page"] = _pid
            if _pid == "results" and st.session_state.get("run_id"):
                st.session_state["results_run_id"] = st.session_state["run_id"]
            st.rerun()

# Gérer la navigation via query params (SUPPRIMER ou laisser pour compatibilité)
_qp = st.query_params.get("nav", None)
if _qp in [p[0] for p in _nav_items]:
    st.session_state["active_page"] = _qp
    st.query_params.clear()
    st.rerun()


if st.session_state.get("notif_banner"):
    n = st.session_state["notif_banner"]
    if n["type"] == "success":
        st.markdown(f'<div style="background:#edfaf0;border:1px solid #a8d5b5;border-left:4px solid #28a745;border-radius:4px;padding:18px 22px;margin-bottom:20px;color:#14532d;font-size:0.95rem;font-weight:500;">{n["msg"]}</div>', unsafe_allow_html=True)
        st.toast(n["msg"], icon="✅")
    else:
        st.markdown(f'<div style="background:#fdf0f0;border:1px solid #f0b8b8;border-left:4px solid #dc3545;border-radius:4px;padding:18px 22px;margin-bottom:20px;color:#7f1d1d;font-size:0.95rem;font-weight:500;">{n["msg"]}</div>', unsafe_allow_html=True)
        st.toast(n["msg"], icon="❌")
    col1, col2 = st.columns([8, 1])
    with col2:
        if st.button("✕", key="close_notif"):
            del st.session_state["notif_banner"]
            _nf = REPO_DIR / "pending_notif.json"
            if _nf.exists():
                _nf.unlink()
            st.rerun()


st.markdown("---")
active_page = st.session_state["active_page"]
active_page = st.session_state["active_page"]


if active_page == "home":
    render_dashboard(load_history, get_run_paths)
# ════════════════════════════════════════════════════════════════════════════
# PAGE : HISTORIQUE
# ════════════════════════════════════════════════════════════════════════════
if active_page == "history":
    st.markdown('<div class="section-label">Historique des analyses</div>', unsafe_allow_html=True)
    history = load_history()
    if not history:
        st.info("Aucune analyse enregistrée. Lancez votre premier pipeline depuis l'onglet **Nouveau pipeline**.")
        st.stop()

    total      = len(history)
    successes  = sum(1 for h in history if h["status"] == "success")
    total_samp = sum(h.get("samples", 0) for h in history)
    durations  = [h["duration_sec"] for h in history if h.get("duration_sec") is not None and h["duration_sec"] > 0]
    avg_dur    = int(sum(durations) / len(durations)) if durations else 0
    m1, m2, m3, m4 = st.columns(4)
    for col, label, value in [
        (m1, "Runs totaux",          str(total)),
        (m2, "Taux de succès",       f"{round(successes/total*100)}%" if total else "—"),
        (m3, "Échantillons traités", str(total_samp)),
        (m4, "Durée moyenne",        f"{avg_dur//60} min {avg_dur%60} s" if avg_dur else "—"),
    ]:
        col.markdown(f'<div class="qc-metric" style="margin-bottom:16px;"><div class="qc-label">{label}</div><div class="qc-value">{value}</div></div>', unsafe_allow_html=True)

    st.markdown("---")
    filter_col, _ = st.columns([2, 3])
    with filter_col:
        status_filter = st.selectbox("Filtrer par statut", ["Tous", "Succès", "Échec"], label_visibility="collapsed")
    status_map = {"Tous": None, "Succès": "success", "Échec": "failed"}
    filtered   = [h for h in history if status_map[status_filter] is None or h["status"] == status_map[status_filter]]

    for h in filtered:
        run_id   = h["run_id"]
        status   = h["status"]
        dur_s    = h.get("duration_sec")
        dur_str  = f"{dur_s//60} min {dur_s%60} s" if dur_s else "—"
        zip_path = Path(h["zip_path"]) if h.get("zip_path") else None
        badge_cls = {"success": "badge-success", "failed": "badge-failed"}.get(status, "badge-running")
        badge_lbl = {"success": "✓ Succès", "failed": "✗ Échec"}.get(status, "⟳ En cours")
        with st.container():
            st.markdown(f"""
            <div class="hist-card">
                <div style="flex:1;">
                    <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">
                        <span class="hist-badge {badge_cls}">{badge_lbl}</span>
                        <span style="font-family:'JetBrains Mono',monospace;font-size:0.78rem;color:#888888;">{run_id}</span>
                    </div>
                    <div style="font-size:0.88rem;color:#888888;">
                        📅 {h.get('date','—')} &nbsp;·&nbsp; 🧬 {h.get('samples','?')} échantillon(s) &nbsp;·&nbsp; ⏱ {dur_str}
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            btn_cols = st.columns([1, 1, 1, 4])
            with btn_cols[0]:
                if zip_path and zip_path.exists():
                    with open(zip_path, "rb") as f:
                        st.download_button("💾", data=f, file_name=f"resultats_{run_id}.zip", mime="application/zip", key=f"dl_{run_id}", help="Télécharger les résultats")
                else:
                    st.button("💾", key=f"dl_{run_id}", disabled=True, help="ZIP non disponible")
            with btn_cols[1]:
                if st.button("📊", key=f"qc_{run_id}", help="Voir le QC"):
                    st.session_state["prev_page"]   = "history"
                    st.session_state["qc_run_id"]   = run_id
                    st.session_state["active_page"] = "qc_detail"
                    st.rerun()
            with btn_cols[2]:
                if st.session_state.get("confirm_delete") == run_id:
                    if st.button("✓ Confirmer", key=f"confirm_{run_id}"):
                        delete_run_from_history(run_id)
                        st.session_state["confirm_delete"] = None
                        st.toast("Run supprimé de l'historique.", icon="🗑️")
                        st.rerun()
                else:
                    if st.button("🗑️", key=f"del_{run_id}", help="Supprimer"):
                        st.session_state["confirm_delete"] = run_id
                        st.rerun()
            if st.session_state.get("confirm_delete") == run_id:
                st.warning(f"Supprimer le run `{run_id}` de l'historique ?")
            sample_list = h.get("sample_list", [])
            if sample_list:
                with st.expander(f"Voir les {len(sample_list)} échantillon(s)"):
                    for s in sample_list:
                        st.code(s, language=None)


# ════════════════════════════════════════════════════════════════════════════
# PAGE : QC DÉTAIL
# ════════════════════════════════════════════════════════════════════════════
elif active_page == "qc_detail":
             render_qc_dashboard(
                 get_qc_data_fn   = get_qc_data_for_run,
                 get_run_paths_fn = get_run_paths,
                 get_badge_fn     = get_module_badge,
             )

# ════════════════════════════════════════════════════════════════════════════
# PAGE : RÉSULTATS MOLÉCULAIRES
# ════════════════════════════════════════════════════════════════════════════
elif active_page == "results":
    render_results_page(get_run_paths)

# ════════════════════════════════════════════════════════════════════════════
# PAGE : PARAMÈTRES
# ════════════════════════════════════════════════════════════════════════════
elif active_page == "settings":
    st.markdown('<div class="section-label">Paramètres</div>', unsafe_allow_html=True)
    with st.expander("🔔 Notifications", expanded=True):
        st.session_state["notif_enabled"] = st.toggle("Activer les notifications", value=st.session_state.get("notif_enabled", True))
        st.session_state["notif_failure_only"] = st.toggle("Alerter seulement en cas d'échec", value=st.session_state.get("notif_failure_only", False))
    with st.expander("🗂️ Maintenance"):
        history  = load_history()
        runs_dir = REPO_DIR / "runs"
        disk_usage = "—"
        if runs_dir.exists():
            try:
                result = subprocess.run(["du", "-sh", str(runs_dir)], capture_output=True, text=True)
                disk_usage = result.stdout.split()[0]
            except Exception:
                pass
        col_a, col_b = st.columns(2)
        col_a.metric("Runs en historique", len(history))
        col_b.metric("Espace utilisé", disk_usage)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Vider l'historique complet", width='stretch'):
            HISTORY_FILE.write_text("[]")
            st.toast("Historique vidé.", icon="🗑️")
            st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# PAGE : PIPELINE
# ════════════════════════════════════════════════════════════════════════════
elif active_page == "pipeline":

    upload_done       = st.session_state.get("upload_done", False)
    nomenclature_done = st.session_state.get("nomenclature_done", False)
    running           = st.session_state.get("running", False)
    pipeline_done     = st.session_state.get("pipeline_done", False)

    def wf_class(step):
        if step == 1:
            return "done" if upload_done else "active"
        if step == 2:
            if not upload_done: return ""
            return "done" if nomenclature_done else "active"
        if step == 3:
            if not nomenclature_done: return ""
            if running: return "active"
            return "done" if pipeline_done else "active"
        if step == 4:
            return "active" if pipeline_done else ""
        return ""

    steps   = [("1","Upload FASTQ"), ("2","Nomenclature AMD"), ("3","Analyse"), ("4","Résultats")]
    wf_cols = st.columns(len(steps))
    for col, (num, label) in zip(wf_cols, steps):
        with col:
            st.markdown(f'<div class="wf-step {wf_class(int(num))}"><div class="wf-num">{num}</div><div class="wf-label">{label}</div></div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # ════════════════════════════════════════════
    # ÉTAPE 1 — UPLOAD
    # ════════════════════════════════════════════
    if not upload_done and not running and not pipeline_done:
        st.markdown('<div class="section-label">Étape 1 — Upload des fichiers FASTQ</div>', unsafe_allow_html=True)
        st.markdown("""
        <div style="background:#fafafa;border:1px solid #d8d8d8;border-radius:4px;padding:14px 18px;margin-bottom:16px;font-size:0.88rem;color:#555555;line-height:1.7;">
            📋 Importez <strong style="color:#6c757d;">tous vos fichiers FASTQ</strong> en une seule fois.<br>
            Les paires R1/R2 sont détectées automatiquement.
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Glissez-déposez vos fichiers FASTQ (.fastq.gz)",
            type=["fastq.gz"],
            accept_multiple_files=True,
            key="fastq_uploader_step1"
        )

        if uploaded:
            r1_files = sorted([f for f in uploaded if re.search(r'[_.]R1[_.]|_1\.fastq', f.name, re.IGNORECASE)], key=lambda x: x.name)
            r2_files = sorted([f for f in uploaded if re.search(r'[_.]R2[_.]|_2\.fastq', f.name, re.IGNORECASE)], key=lambda x: x.name)
            orphans  = [f for f in uploaded if f not in r1_files and f not in r2_files]

            st.markdown(f"""
            <div style="display:flex;gap:12px;margin:16px 0;">
                <div style="background:#ffffff;border:1px solid #dee2e6;border-radius:10px;padding:12px 18px;flex:1;text-align:center;">
                    <div style="font-size:1.5rem;font-weight:700;color:#2c6fad;font-family:'JetBrains Mono',monospace;">{len(r1_files)}</div>
                    <div style="font-size:0.72rem;color:#6c757d;text-transform:uppercase;letter-spacing:1px;margin-top:4px;">Fichiers R1</div>
                </div>
                <div style="background:#ffffff;border:1px solid #dee2e6;border-radius:10px;padding:12px 18px;flex:1;text-align:center;">
                    <div style="font-size:1.5rem;font-weight:700;color:#6f42c1;font-family:'JetBrains Mono',monospace;">{len(r2_files)}</div>
                    <div style="font-size:0.72rem;color:#6c757d;text-transform:uppercase;letter-spacing:1px;margin-top:4px;">Fichiers R2</div>
                </div>
                <div style="background:#ffffff;border:1px solid {"#22c55e33" if len(r1_files)==len(r2_files) and len(r1_files)>0 else "#ef444433"};border-radius:10px;padding:12px 18px;flex:1;text-align:center;">
                    <div style="font-size:1.5rem;font-weight:700;color:{"#4ade80" if len(r1_files)==len(r2_files) and len(r1_files)>0 else "#f87171"};font-family:'JetBrains Mono',monospace;">{min(len(r1_files),len(r2_files))}</div>
                    <div style="font-size:0.72rem;color:#6c757d;text-transform:uppercase;letter-spacing:1px;margin-top:4px;">Paires détectées</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if orphans:
                st.warning(f"⚠️ {len(orphans)} fichier(s) non reconnu(s) comme R1 ou R2 (ignorés) : " + ", ".join(f.name for f in orphans))

            if len(r1_files) == 0:
                st.error("❌ Aucun fichier R1 détecté.")
            elif len(r1_files) != len(r2_files):
                st.error(f"❌ Déséquilibre R1/R2 : {len(r1_files)} R1 et {len(r2_files)} R2.")
            else:
                # ── Appairage par préfixe ──────────────────────────────────────────
                r1_by_prefix = {get_pair_prefix(f.name): f for f in r1_files}
                r2_by_prefix = {get_pair_prefix(f.name): f for f in r2_files}

                matched_pairs  = []   # [(f1, f2), ...]
                mismatched_r1  = []   # R1 sans R2 correspondant
                mismatched_r2  = []   # R2 sans R1 correspondant

                all_prefixes = sorted(set(r1_by_prefix) | set(r2_by_prefix))
                for prefix in all_prefixes:
                    has_r1 = prefix in r1_by_prefix
                    has_r2 = prefix in r2_by_prefix
                    if has_r1 and has_r2:
                        matched_pairs.append((r1_by_prefix[prefix], r2_by_prefix[prefix]))
                    elif has_r1:
                        mismatched_r1.append(r1_by_prefix[prefix])
                    else:
                        mismatched_r2.append(r2_by_prefix[prefix])

                # ── Signalement des fichiers sans correspondance ───────────────────
                if mismatched_r1:
                    for f in mismatched_r1:
                        st.error(
                            f"❌ **R1 sans R2 correspondant** : `{f.name}`  \n"
                            f"Aucun fichier R2 avec le même préfixe n'a été trouvé."
                        )
                if mismatched_r2:
                    for f in mismatched_r2:
                        st.error(
                            f"❌ **R2 sans R1 correspondant** : `{f.name}`  \n"
                            f"Aucun fichier R1 avec le même préfixe n'a été trouvé."
                        )

                has_pair_errors = bool(mismatched_r1 or mismatched_r2)

                if not matched_pairs and has_pair_errors:
                    st.stop()

                if has_pair_errors:
                    st.warning(
                        f"⚠️ Seules les {len(matched_pairs)} paires correctement appariées "
                        f"sont affichées ci-dessous. Corrigez les fichiers en erreur avant de continuer."
                    )

                # ════════════════════════════════════════════════════════════
                # CHANGE 1 — Tableau récapitulatif : ID Sample | Forward | Reverse
                # ════════════════════════════════════════════════════════════
                st.markdown('<div class="section-label" style="margin-top:16px;">Récapitulatif des paires détectées</div>', unsafe_allow_html=True)
                total_size = sum(f.size for f in [f for f1, f2 in matched_pairs for f in (f1, f2)])
                size_str   = f"{total_size/1e9:.2f} Go" if total_size > 1e9 else f"{total_size/1e6:.1f} Mo"
                st.markdown(f'<div style="font-size:0.8rem;color:#888888;margin-bottom:10px;">📦 {len(matched_pairs)} paire(s) · {size_str} au total</div>', unsafe_allow_html=True)

                table_rows_html = ""

                for i, (f1, f2) in enumerate(matched_pairs):
                    sample_id = get_sample_id(f1.name)
                    size_info = f"{round(f1.size/1e6,1)}+{round(f2.size/1e6,1)} Mo"

                    table_rows_html += (
                        f"<tr>"
                        f"<td class='col-idx'>{i+1}</td>"
                        f"<td class='col-id'>{sample_id}</td>"
                        f"<td class='col-fwd'>{f1.name}</td>"
                        f"<td class='col-rev'>{f2.name}</td>"
                        f"<td style='color:#aaaaaa;font-size:0.72rem;white-space:nowrap;'>{size_info}</td>"
                        f"</tr>"
                    )

                table_html = f"""
                <div style="overflow-x:auto;">
                <table class="upload-table">
                    <thead>
                        <tr>
                            <th style="width:32px;">#</th>
                            <th class="col-id">ID Sample</th>
                            <th class="col-fwd">Forward (R1)</th>
                            <th class="col-rev">Reverse (R2)</th>
                            <th style="background:#2a3a5a;">Taille</th>
                        </tr>
                    </thead>
                    <tbody>
                        {table_rows_html}
                    </tbody>
                </table>
                </div>
                """

                st.markdown(table_html, unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)
                #pairs_built = [{"r1": f1, "r2": f2} for f1, f2 in zip(r1_files, r2_files)]
                pairs_built = [{"r1": f1, "r2": f2} for f1, f2 in matched_pairs]

                if not has_pair_errors:
                    col_direct, col_rename = st.columns(2)
                    with col_direct:
                        if st.button("▶  Lancer directement\n\nMes fichiers sont déjà au format AMD", width='stretch'):
                            sample_names = []
                            for p in pairs_built:
                                base = re.sub(r'[_.]R[12].*$', '', p["r1"].name, flags=re.IGNORECASE)
                                sample_names.append({
                                    "base": base, "type": "Direct",
                                    "r1": p["r1"].name, "r2": p["r2"].name,
                                    "orig_r1": p["r1"].name, "orig_r2": p["r2"].name,
                                    "file_r1": p["r1"], "file_r2": p["r2"],
                                })
                            st.session_state["uploaded_pairs"]    = pairs_built
                            st.session_state["upload_done"]       = True
                            st.session_state["naming_mode"]       = "direct"
                            st.session_state["sample_names"]      = sample_names
                            st.session_state["nomenclature_done"] = True
                            st.rerun()
                    with col_rename:
                        if st.button("🏷️  Renommer les échantillons\n\nAssocier chaque fichier à sa nomenclature AMD", width='stretch'):
                            st.session_state["uploaded_pairs"] = pairs_built
                            st.session_state["upload_done"]    = True
                            st.session_state["naming_mode"]    = None
                            st.rerun()
                else:
                    st.info("ℹ️ Résolvez les erreurs d'appairage avant de pouvoir continuer.")

    # ════════════════════════════════════════════
    # ÉTAPE 2 — NOMENCLATURE AMD
    # ════════════════════════════════════════════
    elif upload_done and not nomenclature_done and not running and not pipeline_done:

        pairs = st.session_state.get("uploaded_pairs", [])

        if st.button("← Retour à l'upload"):
            st.session_state["upload_done"]   = False
            st.session_state["naming_mode"]   = None
            st.session_state["sample_names"]  = []
            st.rerun()

        st.markdown('<div class="section-label">Étape 2 — Nomenclature AMD</div>', unsafe_allow_html=True)
        st.markdown(f'<div style="background:#fafafa;border:1px solid #d8d8d8;border-radius:4px;padding:12px 16px;margin-bottom:16px;font-size:0.88rem;color:#555555;">📋 <strong style="color:#6c757d;">{len(pairs)} paire(s)</strong> uploadée(s) — choisissez comment les nommer.</div>', unsafe_allow_html=True)

        naming_mode = st.session_state.get("naming_mode")

        if naming_mode is None:
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("📝  Tableau de saisie manuelle\n\nRemplir les champs AMD ligne par ligne", width='stretch'):
                    st.session_state["naming_mode"] = "manual"
                    st.rerun()
            with col_b:
                if st.button("📊  Importer depuis Excel\n\nRemplir les champs AMD via un fichier Excel", width='stretch'):
                    st.session_state["naming_mode"] = "excel"
                    st.rerun()
            st.stop()

        if st.button("← Changer de mode"):
            st.session_state["naming_mode"] = None
            st.rerun()

        # ════════════════════════════════
        # MODE EXCEL
        # ════════════════════════════════
        if naming_mode == "excel":
            st.markdown('<div class="section-label">Import Excel</div>', unsafe_allow_html=True)

            # Mini-tableau de référence des paires
            table_rows_html = ""
            for i, pair in enumerate(pairs):
                sample_id = get_sample_id(pair["r1"].name)
                table_rows_html += f"""
                <tr>
                    <td class="col-idx">{i+1}</td>
                    <td class="col-id">{sample_id}</td>
                    <td class="col-fwd">{pair['r1'].name}</td>
                    <td class="col-rev">{pair['r2'].name}</td>
                </tr>"""
            st.markdown(f"""
            <div style="overflow-x:auto;margin-bottom:14px;">
            <table class="upload-table">
                <thead>
                    <tr>
                        <th style="width:32px;">#</th>
                        <th class="col-id">ID Sample</th>
                        <th class="col-fwd">Forward (R1)</th>
                        <th class="col-rev">Reverse (R2)</th>
                    </tr>
                </thead>
                <tbody>{table_rows_html}</tbody>
            </table>
            </div>
            """, unsafe_allow_html=True)

            st.download_button(
                "⬇️ Télécharger le template Excel",
                data=generate_template(pairs),          # ← passer pairs ici
                file_name="template_nomenclature_AMD.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
                )

            st.markdown(f"""
            <div style="background:#e8f2fb;border:1px solid #bdd5ef;border-radius:4px;padding:10px 14px;margin:12px 0;font-size:0.85rem;color:#1a3a6a;">
                ℹ️ Le fichier Excel doit contenir <strong>{len(pairs)} ligne(s)</strong> dans le même ordre que les paires listées ci-dessus.<br>
                La première colonne <strong>ID Sample</strong> identifie la paire (partie avant le '_' du nom de fichier).<br>
                ⚠️ Ne pas modifier la colonne <strong>ID Sample</strong> — elle sert à identifier vos fichiers.<br>
                Colonnes mol markers : une colonne par marqueur avec <strong>0</strong> ou <strong>1</strong>.<br>
                Le code décimal est calculé automatiquement (ex : k13+crt = 110000000 = 384).
            </div>
            """, unsafe_allow_html=True)

            excel_file = st.file_uploader("Importez votre fichier Excel (.xlsx)", type=["xlsx"], key="excel_nomenclature")
            
            if excel_file:
                try:
                    wb   = openpyxl.load_workbook(excel_file)
                    ws   = wb.active
                    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]
                    errors = []; names = []
 
                    # Construire un index pairs : id_sample → pair
                    # get_sample_id extrait la partie avant le premier '_'
                    pairs_by_id = {}
                    for pair in pairs:
                        sid = get_sample_id(pair["r1"].name)
                        pairs_by_id[sid] = pair
 
                    for row_idx, row in enumerate(rows, start=2):
                        cells = [str(v).strip() if v is not None else "" for v in row]
                        while len(cells) < 20:
                            cells.append("")
 
                        id_sample   = cells[0].strip()
                        sample_type = cells[1].upper()
                        year        = cells[2].upper()
                        country     = cells[3].upper()
                        state       = cells[4].upper()
                        day         = cells[5].upper()
                        treat_raw   = cells[6].upper()
                        sid_ind     = cells[7].upper()
                        sid3_pool   = cells[8].upper()
                        npool       = cells[9].upper()
                        mol_bits    = cells[10:19]
                        proc_raw    = cells[19].strip()
 
                        # ── Matching par ID Sample ──────────────────────────
                        pair = pairs_by_id.get(id_sample)
                        if pair is None:
                            # Tentative de matching insensible à la casse
                            pair = next(
                                (p for k, p in pairs_by_id.items()
                                 if k.lower() == id_sample.lower()),
                                None
                            )
                        if pair is None:
                            errors.append(
                                f"Ligne {row_idx} : ID Sample « {id_sample} » "
                                f"ne correspond à aucun fichier uploadé. "
                                f"IDs disponibles : {', '.join(pairs_by_id.keys())}"
                            )
                            continue
 
                        # ── Processé ───────────────────────────────────────
                        if proc_raw in ("1", "2"):
                            proc = proc_raw
                        elif proc_raw.upper() == "Z":
                            proc = "1"
                        elif proc_raw.upper() == "Y":
                            proc = "2"
                        else:
                            proc = proc_raw.upper()
 
                        # ── Code traitement ────────────────────────────────
                        drug_match = next((k for k in DRUG_CODES if k.lower() == treat_raw.lower()), None)
                        if drug_match:
                            treat = DRUG_CODES[drug_match]
                        else:
                            treat = treat_raw
 
                        # ── Mol markers ────────────────────────────────────
                        selected_markers = [
                            MOL_MARKERS_LIST[i]
                            for i, b in enumerate(mol_bits)
                            if str(b).strip() in ("1", "1.0", "True", "true", "oui", "yes")
                        ]
                        mol = compute_mol_code(selected_markers)
 
                        # ── Validation ─────────────────────────────────────
                        if sample_type not in ("INDIVIDUAL", "POOLED"):
                            errors.append(f"Ligne {row_idx} ({id_sample}) : Type doit être 'Individual' ou 'Pooled'")
                            continue
 
                        row_errors = validate_common(year, country, state, day, treat, mol, proc)
                        for e in row_errors:
                            errors.append(f"Ligne {row_idx} ({id_sample}) : {e}")
 
                        if sample_type == "INDIVIDUAL":
                            if not only_digits(sid_ind, 4):
                                errors.append(f"Ligne {row_idx} ({id_sample}) : Sample ID Individual doit être 4 chiffres")
                            if not row_errors and only_digits(sid_ind, 4):
                                base    = build_base(year, country, state, day, treat, sid_ind, "Pf" + "B" + mol + proc)
                                r1_name = build_new_filename(base, pair["r1"].name)
                                r2_name = build_new_filename(base, pair["r2"].name)
                                names.append({
                                    "base": base, "type": "Individual",
                                    "r1": r1_name, "r2": r2_name,
                                    "orig_r1": pair["r1"].name, "orig_r2": pair["r2"].name,
                                    "file_r1": pair["r1"], "file_r2": pair["r2"],
                                    "mol_markers": selected_markers, "mol_code": mol,
                                    "drug": drug_match if drug_match else "",
                                })
                        else:  # POOLED
                            if not only_digits(sid3_pool, 3):
                                errors.append(f"Ligne {row_idx} ({id_sample}) : Sample ID Pooled doit être 3 chiffres")
                            if not only_digits(npool, 2):
                                errors.append(f"Ligne {row_idx} ({id_sample}) : Nb dans pool doit être 2 chiffres")
                            if not row_errors and only_digits(sid3_pool, 3) and only_digits(npool, 2):
                                base    = build_base(year, country, state, day, treat, sid3_pool + "P" + npool, "B" + mol + proc)
                                r1_name = build_new_filename(base, pair["r1"].name)
                                r2_name = build_new_filename(base, pair["r2"].name)
                                names.append({
                                    "base": base, "type": "Pooled",
                                    "r1": r1_name, "r2": r2_name,
                                    "orig_r1": pair["r1"].name, "orig_r2": pair["r2"].name,
                                    "file_r1": pair["r1"], "file_r2": pair["r2"],
                                    "mol_markers": selected_markers, "mol_code": mol,
                                    "drug": drug_match if drug_match else "",
                                })
 
                    # Vérifier que tous les fichiers uploadés ont une ligne dans l'Excel
                    matched_ids = set()
                    for row in rows:
                        cells = [str(v).strip() if v is not None else "" for v in row]
                        if cells:
                            matched_ids.add(cells[0].strip())
                    unmatched = [sid for sid in pairs_by_id if sid not in matched_ids
                                 and sid.lower() not in {m.lower() for m in matched_ids}]
                    if unmatched:
                        st.warning(
                            f"⚠️ {len(unmatched)} fichier(s) uploadé(s) sans ligne dans l'Excel : "
                            + ", ".join(unmatched)
                        )
 
                    if errors:
                        st.error(f"❌ {len(errors)} erreur(s) :")
                        for e in errors:
                            st.markdown(f"- {e}")
                    elif not names:
                        st.warning("⚠️ Aucun échantillon valide trouvé.")
                    else:
                        st.markdown('<div class="section-label">Correspondance fichiers → nomenclature</div>', unsafe_allow_html=True)
                        for i, n in enumerate(names):
                            badge = "🔵" if n["type"] == "Individual" else "🟢"
                            st.markdown(f"""
                            <div class="assoc-row">
                                <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;">
                                    <span style="font-size:0.7rem;font-weight:700;color:#6c757d;">#{i+1}</span>
                                    <span style="font-size:0.7rem;font-weight:700;color:#888888;text-transform:uppercase;">{badge} {n['type']}</span>
                                    <span style="font-size:0.72rem;color:#7c3aed;background:#faf5ff;border:1px solid #d8b4fe;border-radius:3px;padding:1px 7px;">
                                        🧬 mol: {n['mol_code']} ({", ".join(n['mol_markers']) if n['mol_markers'] else "aucun"})
                                    </span>
                                </div>
                                <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;">
                                    <span class="assoc-orig">{n['orig_r1']}</span>
                                    <span style="color:#2c6fad;">→</span>
                                    <span class="assoc-name">{n['r1']}</span>
                                </div>
                                <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-top:4px;">
                                    <span class="assoc-orig">{n['orig_r2']}</span>
                                    <span style="color:#6f42c1;">→</span>
                                    <span class="assoc-name" style="color:#6f42c1;">{n['r2']}</span>
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                            render_mol_bitcode(n["mol_markers"], n["mol_code"])
 
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("✅ Valider la nomenclature et continuer", width='stretch'):
                            st.session_state["sample_names"]      = names
                            st.session_state["nomenclature_done"] = True
                            st.rerun()
                except Exception as e:
                    st.error(f"❌ Erreur lors de la lecture : {e}")


        # ════════════════════════════════════════════════════════════════════
        # MODE MANUEL — CHANGE 2: tableau compact + "appliquer à tous"
        # ════════════════════════════════════════════════════════════════════
 
        elif naming_mode == "manual":
            st.markdown('<div class="section-label">Saisie manuelle — Tableau AMD</div>', unsafe_allow_html=True)
 
            n_pairs = len(pairs)
 
            st.markdown("""
            <style>
            .mol-pill {
                display:inline-block; font-size:0.7rem; background:#faf5ff;
                border:1px solid #d8b4fe; border-radius:10px; padding:2px 7px;
                color:#7c3aed; font-family:'JetBrains Mono',monospace; font-weight:600; margin-top:2px;
            }
            .fill-banner {
                background:#e8f2fb; border:1px solid #bdd5ef;
                border-radius:6px; padding:10px 14px; margin:8px 0 14px;
            }
            .preview-r1 { color:#2c6fad; font-family:'JetBrains Mono',monospace; font-size:0.78rem; font-weight:600; }
            .preview-r2 { color:#6f42c1; font-family:'JetBrains Mono',monospace; font-size:0.78rem; font-weight:600; margin-top:2px; }
            </style>
            """, unsafe_allow_html=True)
 
            # ── Initialisation ───────────────────────────────────────────────
            if "manual_rows" not in st.session_state or len(st.session_state["manual_rows"]) != n_pairs:
                st.session_state["manual_rows"] = [
                    {
                        "year": "", "country": "", "state": "", "day": "",
                        "treat": "", "type": "Individual",
                        "sid": "", "sid3": "", "npool": "",
                        "mol_markers": [], "proc": "1", "drug": DRUG_OPTIONS[0],
                    }
                    for _ in pairs
                ]
            rows_state = st.session_state["manual_rows"]
 
            # ── Compteur de rerun pour rendre les clés uniques ───────────────
            if "manual_rerun_count" not in st.session_state:
                st.session_state["manual_rerun_count"] = 0
            rc_count = st.session_state["manual_rerun_count"]
 
            # ── Colonnes ─────────────────────────────────────────────────────
            COL_W      = [0.28, 1.25, 0.85, 0.6, 0.6, 0.6, 0.6, 1.85, 0.55, 0.6, 1.05, 1.2]
            COL_LABELS = ["#", "ID Sample", "Type", "Année", "Pays", "État",
                          "Jour", "Médicament", "Code", "Proc.", "SID / Pool", "Mol markers"]
            HEADER_BG  = ["#2a3a5a","#0f4c81","#1a2a4a","#1a2a4a","#1a2a4a","#1a2a4a",
                          "#1a2a4a","#1a2a4a","#1a2a4a","#1a2a4a","#1a2a4a","#4a235a"]
 
            hdr = st.columns(COL_W)
            for col, label, bg in zip(hdr, COL_LABELS, HEADER_BG):
                col.markdown(
                    f'<div style="background:{bg};color:#fff;font-size:0.65rem;font-weight:700;'
                    f'letter-spacing:1.1px;text-transform:uppercase;padding:8px 4px;'
                    f'text-align:center;white-space:nowrap;">{label}</div>',
                    unsafe_allow_html=True
                )
 
            # ── Registre des clés actives (pour lecture après) ───────────────
            active_keys = {}   # i -> dict de clés
 
            # ════════════════════════════════════════════════════════════════
            # LIGNES
            # La clé de chaque widget = f"{prefix}_{i}_{rc_count}_{valeur}"
            # → si valeur change (fill-all), clé change → nouveau widget
            # ════════════════════════════════════════════════════════════════
            for i, (pair, rv) in enumerate(zip(pairs, rows_state)):
                sample_id = get_sample_id(pair["r1"].name)
                rc = st.columns(COL_W)
                active_keys[i] = {}
 
                # Col 0 — Numéro
                rc[0].markdown(
                    f'<div style="padding:10px 2px;text-align:center;font-size:0.72rem;'
                    f'color:#aaa;font-weight:600;border-top:1px solid #eee;">{i+1}</div>',
                    unsafe_allow_html=True
                )
 
                # Col 1 — ID Sample (lecture seule)
                rc[1].markdown(
                    f'<div style="background:#f0f4ff;padding:10px 8px;border-top:1px solid #eee;'
                    f'border-right:2px solid #2c6fad;font-family:\'JetBrains Mono\',monospace;'
                    f'font-size:0.8rem;font-weight:700;color:#1a2a4a;white-space:nowrap;'
                    f'overflow:hidden;text-overflow:ellipsis;" title="{sample_id}">{sample_id}</div>',
                    unsafe_allow_html=True
                )
 
                # Col 2 — Type
                # La clé encode la valeur courante → si fill-all change type, nouvelle clé
                k_type = f"w_type_{i}_{rc_count}"
                rc[2].radio(
                    f"Type {i+1}", ["Ind.", "Pool"],
                    key=k_type,
                    index=0 if rv["type"] == "Individual" else 1,
                    horizontal=True, label_visibility="collapsed"
                )
                active_keys[i]["type"] = k_type
 
                # Col 3 — Année
                k_year = f"w_year_{i}_{rc_count}"
                rc[3].text_input(
                    f"Année {i+1}", max_chars=2, placeholder="24",
                    key=k_year, value=rv["year"],
                    label_visibility="collapsed"
                )
                active_keys[i]["year"] = k_year
 
                # Col 4 — Pays
                k_country = f"w_country_{i}_{rc_count}"
                rc[4].text_input(
                    f"Pays {i+1}", max_chars=2, placeholder="SN",
                    key=k_country, value=rv["country"],
                    label_visibility="collapsed"
                )
                active_keys[i]["country"] = k_country
 
                # Col 5 — État
                k_state = f"w_state_{i}_{rc_count}"
                rc[5].text_input(
                    f"État {i+1}", max_chars=2, placeholder="DK",
                    key=k_state, value=rv["state"],
                    label_visibility="collapsed"
                )
                active_keys[i]["state"] = k_state
 
                # Col 6 — Jour
                k_day = f"w_day_{i}_{rc_count}"
                rc[6].text_input(
                    f"Jour {i+1}", max_chars=2, placeholder="00",
                    key=k_day, value=rv["day"],
                    label_visibility="collapsed"
                )
                active_keys[i]["day"] = k_day
 
                # Col 7 — Médicament
                drug_idx  = DRUG_OPTIONS.index(rv["drug"]) if rv["drug"] in DRUG_OPTIONS else 0
                k_drug    = f"w_drug_{i}_{rc_count}"
                rc[7].selectbox(
                    f"Méd {i+1}", DRUG_OPTIONS, index=drug_idx,
                    key=k_drug, label_visibility="collapsed"
                )
                active_keys[i]["drug"] = k_drug
 
                # Col 8 — Code traitement
                cur_drug = st.session_state.get(k_drug, rv["drug"])
                if cur_drug != DRUG_OPTIONS[0]:
                    rc[8].markdown(
                        f'<div style="background:#edfaf0;border:1px solid #a8d5b5;border-radius:3px;'
                        f'padding:7px 6px;font-family:\'JetBrains Mono\',monospace;font-size:0.9rem;'
                        f'font-weight:700;color:#155724;text-align:center;">'
                        f'{DRUG_CODES[cur_drug]}</div>',
                        unsafe_allow_html=True
                    )
                    active_keys[i]["treat"] = None  # calculé depuis drug
                else:
                    k_treat = f"w_treat_{i}_{rc_count}"
                    rc[8].text_input(
                        f"Code {i+1}", max_chars=1, placeholder="A",
                        key=k_treat, value=rv.get("treat", ""),
                        label_visibility="collapsed"
                    )
                    active_keys[i]["treat"] = k_treat
 
                # Col 9 — Processé
                k_proc = f"w_proc_{i}_{rc_count}"
                rc[9].radio(
                    f"Proc {i+1}", ["1", "2"],
                    key=k_proc,
                    index=0 if rv["proc"] != "2" else 1,
                    horizontal=True, label_visibility="collapsed"
                )
                active_keys[i]["proc"] = k_proc
 
                # Col 10 — SID / Pool
                cur_type_raw = st.session_state.get(k_type, "Ind.")
                cur_type_str = "Individual" if cur_type_raw == "Ind." else "Pooled"
                if cur_type_str == "Individual":
                    k_sid = f"w_sid_{i}_{rc_count}"
                    rc[10].text_input(
                        f"SID {i+1}", max_chars=4, placeholder="1000",
                        key=k_sid, value=rv.get("sid", ""),
                        label_visibility="collapsed"
                    )
                    active_keys[i]["sid"]   = k_sid
                    active_keys[i]["sid3"]  = None
                    active_keys[i]["npool"] = None
                else:
                    c_a, c_sep, c_b = rc[10].columns([2, 0.4, 2])
                    k_sid3 = f"w_sid3_{i}_{rc_count}"
                    c_a.text_input(
                        f"SID3 {i+1}", max_chars=3, placeholder="000",
                        key=k_sid3, value=rv.get("sid3", ""),
                        label_visibility="collapsed"
                    )
                    c_sep.markdown(
                        '<div style="text-align:center;font-weight:700;padding-top:8px;color:#888;">P</div>',
                        unsafe_allow_html=True
                    )
                    k_npool = f"w_npool_{i}_{rc_count}"
                    c_b.text_input(
                        f"Npool {i+1}", max_chars=2, placeholder="05",
                        key=k_npool, value=rv.get("npool", ""),
                        label_visibility="collapsed"
                    )
                    active_keys[i]["sid"]   = None
                    active_keys[i]["sid3"]  = k_sid3
                    active_keys[i]["npool"] = k_npool
 
                # Col 11 — Mol markers
                # La clé encode rc_count + les marqueurs actuels pour forcer
                # la recréation du widget après fill-all (même logique que text_input)
                k_mol = f"w_mol_{i}_{rc_count}"
                rc[11].multiselect(
                    f"Mol {i+1}", options=MOL_MARKERS_LIST,
                    default=rv.get("mol_markers", []),
                    key=k_mol,
                    label_visibility="collapsed",
                    placeholder="Sélectionner…"
                )
                active_keys[i]["mol"] = k_mol
                mol_code = compute_mol_code(rv.get("mol_markers", []))
                if rv.get("mol_markers"):
                    rc[11].markdown(
                        f'<div class="mol-pill">{mol_code} ({len(rv["mol_markers"])})</div>',
                        unsafe_allow_html=True
                    )
 
            # ════════════════════════════════════════════════════════════════
            # LECTURE session_state → rows_state (après TOUS les widgets)
            # ════════════════════════════════════════════════════════════════
            for i, rv in enumerate(rows_state):
                keys = active_keys[i]
 
                raw_type = st.session_state.get(keys["type"], "Ind.")
                rv["type"] = "Individual" if raw_type == "Ind." else "Pooled"
 
                rv["year"]    = st.session_state.get(keys["year"],    rv["year"])
                rv["country"] = st.session_state.get(keys["country"], rv["country"])
                rv["state"]   = st.session_state.get(keys["state"],   rv["state"])
                rv["day"]     = st.session_state.get(keys["day"],     rv["day"])
                rv["proc"]    = st.session_state.get(keys["proc"],    rv["proc"])
 
                raw_drug = st.session_state.get(keys["drug"], rv["drug"])
                rv["drug"] = raw_drug
                if raw_drug != DRUG_OPTIONS[0]:
                    rv["treat"] = DRUG_CODES[raw_drug]
                elif keys["treat"]:
                    rv["treat"] = st.session_state.get(keys["treat"], rv.get("treat", ""))
 
                if rv["type"] == "Individual":
                    rv["sid"]   = st.session_state.get(keys["sid"],   rv.get("sid", "")) if keys["sid"] else ""
                    rv["sid3"]  = ""
                    rv["npool"] = ""
                else:
                    rv["sid3"]  = st.session_state.get(keys["sid3"],  rv.get("sid3", ""))  if keys["sid3"]  else ""
                    rv["npool"] = st.session_state.get(keys["npool"], rv.get("npool", "")) if keys["npool"] else ""
                    rv["sid"]   = ""
 
                rv["mol_markers"] = list(st.session_state.get(keys["mol"], rv.get("mol_markers", [])))
 
            # ════════════════════════════════════════════════════════════════
            # BANDEAU "REMPLIR POUR TOUS"
            # Au clic : met à jour rows_state + incrémente rc_count
            # → toutes les clés changent → nouveaux widgets avec value= propagée
            # ════════════════════════════════════════════════════════════════
            if n_pairs > 1:
 
                PROP = [
                    ("type",        "Type",           lambda rv: rv["type"]),
                    ("year",        "Année",           lambda rv: rv["year"]),
                    ("country",     "Pays",            lambda rv: rv["country"]),
                    ("state",       "État",            lambda rv: rv["state"]),
                    ("day",         "Jour",            lambda rv: rv["day"]),
                    ("drug",        "Médicament",      lambda rv: rv["drug"]),
                    ("proc",        "Processé",        lambda rv: rv["proc"]),
                    ("sid",         "Sample ID Ind.",  lambda rv: rv.get("sid", "")),
                    ("sid3",        "ID Pooled",       lambda rv: rv.get("sid3", "")),
                    ("npool",       "Nb pool",         lambda rv: rv.get("npool", "")),
                    ("mol_markers", "Mol markers",     lambda rv: rv.get("mol_markers", [])),
                ]
 
                def _first_non_empty(getter):
                    for rv in rows_state:
                        v = getter(rv)
                        if v and v != DRUG_OPTIONS[0] and v != []:
                            return v
                    return None
 
                buttons = []
                for fk, flabel, getter in PROP:
                    ref = _first_non_empty(getter)
                    if ref:
                        buttons.append((fk, flabel, ref))
 
                if buttons:
                    st.markdown(
                        '<div class="fill-banner">'
                        '<div style="font-size:0.75rem;font-weight:700;color:#2c6fad;margin-bottom:8px;">'
                        '↓ Remplir pour tous les autres échantillons :</div>',
                        unsafe_allow_html=True
                    )
                    MAX_PER_ROW = 4
                    for chunk in [buttons[j:j+MAX_PER_ROW] for j in range(0, len(buttons), MAX_PER_ROW)]:
                        bcols = st.columns(len(chunk))
                        for bcol, (fk, flabel, ref_val) in zip(bcols, chunk):
                            disp = ", ".join(ref_val) if isinstance(ref_val, list) else str(ref_val)
                            if len(disp) > 13:
                                disp = disp[:11] + "…"
 
                            if bcol.button(
                                f"Même {flabel}\n`{disp}`",
                                key=f"fillall_{fk}",
                                help=f"Appliquer « {disp} » à tous"
                            ):
                                for rv in rows_state:
                                    if fk == "mol_markers":
                                        rv["mol_markers"] = list(ref_val) if isinstance(ref_val, list) else []
                                    elif fk == "drug":
                                        rv["drug"]  = ref_val
                                        rv["treat"] = DRUG_CODES.get(ref_val, rv.get("treat", ""))
                                    else:
                                        rv[fk] = ref_val
 
                                # Incrémenter rc_count → toutes les clés changent
                                # → Streamlit recrée les widgets avec value= depuis rows_state
                                st.session_state["manual_rerun_count"] += 1
                                st.toast(f"✓ « {flabel} » → {disp} propagé à tous", icon="✅")
                                st.rerun()
 
                    st.markdown('</div>', unsafe_allow_html=True)
 
            st.markdown("---")
 
            # ════════════════════════════════════════════════════════════════
            # PRÉVISUALISATION DES NOMS AMD
            # ════════════════════════════════════════════════════════════════
            st.markdown(
                '<div style="font-size:0.8rem;font-weight:700;color:#28a745;letter-spacing:1px;'
                'text-transform:uppercase;margin-bottom:8px;">Prévisualisation des noms AMD générés</div>',
                unsafe_allow_html=True
            )
 
            names_preview = []
            all_valid     = True
 
            for i, (pair, rv) in enumerate(zip(pairs, rows_state)):
                mol_code   = compute_mol_code(rv.get("mol_markers", []))
                errors_row = validate_common(
                    rv["year"], rv["country"], rv["state"],
                    rv["day"], rv["treat"], mol_code, rv["proc"]
                )
                base = ""
                if not errors_row:
                    if rv["type"] == "Individual":
                        sid = rv.get("sid", "")
                        if only_digits(sid, 4):
                            base = build_base(
                                rv["year"], rv["country"], rv["state"],
                                rv["day"], rv["treat"], sid,
                                "Pf" + "B" + mol_code + rv["proc"]
                            )
                    else:
                        sid3  = rv.get("sid3", "")
                        npool = rv.get("npool", "")
                        if only_digits(sid3, 3) and only_digits(npool, 2):
                            base = build_base(
                                rv["year"], rv["country"], rv["state"],
                                rv["day"], rv["treat"], sid3 + "P" + npool,
                                "B" + mol_code + rv["proc"]
                            )
 
                sample_id = get_sample_id(pair["r1"].name)
                if base:
                    r1_new = build_new_filename(base, pair["r1"].name)
                    r2_new = build_new_filename(base, pair["r2"].name)
                    st.markdown(f"""
                    <div style="background:#f0fdf4;border:1px solid #a8d5b5;border-radius:4px;
                                padding:8px 14px;margin-bottom:5px;">
                        <div style="font-size:0.7rem;color:#888;margin-bottom:3px;
                                    font-family:'JetBrains Mono',monospace;">#{i+1} {sample_id}</div>
                        <div class="preview-r1">{r1_new}</div>
                        <div class="preview-r2">{r2_new}</div>
                    </div>
                    """, unsafe_allow_html=True)
                    names_preview.append({
                        "base": base, "type": rv["type"],
                        "r1": r1_new, "r2": r2_new,
                        "orig_r1": pair["r1"].name, "orig_r2": pair["r2"].name,
                        "file_r1": pair["r1"], "file_r2": pair["r2"],
                        "mol_markers": rv.get("mol_markers", []),
                        "mol_code": mol_code,
                        "drug": rv.get("drug", DRUG_OPTIONS[0]),
                    })
                else:
                    all_valid = False
                    names_preview.append(None)
                    err_msg = " · ".join(errors_row) if errors_row else "Identifiants incomplets"
                    st.markdown(
                        f'<div style="background:#fdf0f0;border:1px solid #f0b8b8;border-radius:4px;'
                        f'padding:7px 14px;margin-bottom:5px;font-size:0.82rem;color:#721c24;">'
                        f'❌ #{i+1} {sample_id} — {err_msg}</div>',
                        unsafe_allow_html=True
                    )
 
            bases = [n["base"] for n in names_preview if n]
            if len(bases) != len(set(bases)):
                st.warning("⚠️ Des noms AMD identiques ont été générés — vérifiez vos identifiants.")
                all_valid = False
 
            st.markdown("<br>", unsafe_allow_html=True)
            if all_valid and len(names_preview) == n_pairs:
                st.success(f"✅ {len(names_preview)} échantillon(s) prêts")
                if st.button("✅ Valider la nomenclature et continuer", width='stretch'):
                    st.session_state["sample_names"]      = names_preview
                    st.session_state["nomenclature_done"] = True
                    st.rerun()
            else:
                st.info("ℹ️ Complétez tous les champs pour activer la validation.")
 

    # ════════════════════════════════════════════
    # RÉCAPITULATIF (après nomenclature validée)
    # ════════════════════════════════════════════
    elif nomenclature_done and not running and not pipeline_done:
        names = st.session_state.get("sample_names", [])

        if st.button("← Modifier la nomenclature"):
            st.session_state["nomenclature_done"] = False
            st.session_state["sample_names"]      = []
            if st.session_state.get("naming_mode") == "direct":
                st.session_state["naming_mode"] = None
            st.rerun()

        st.markdown('<div class="section-label">Récapitulatif — Correspondance fichiers → nomenclature AMD</div>', unsafe_allow_html=True)

        n_ind  = sum(1 for n in names if n["type"] == "Individual")
        n_pool = sum(1 for n in names if n["type"] == "Pooled")
        n_dir  = sum(1 for n in names if n["type"] == "Direct")
        type_summary = []
        if n_ind:  type_summary.append(f"🔵 {n_ind} Individual")
        if n_pool: type_summary.append(f"🟢 {n_pool} Pooled")
        if n_dir:  type_summary.append(f"📂 {n_dir} Direct")

        st.markdown(f'<div style="background:#fafafa;border:1px solid #d8d8d8;border-radius:4px;padding:12px 16px;margin-bottom:14px;font-size:0.9rem;color:#555555;">{len(names)} échantillon(s) · {" · ".join(type_summary)}</div>', unsafe_allow_html=True)

        if "editing_idx" not in st.session_state:
            st.session_state["editing_idx"] = None

        for i, n in enumerate(names):
            badge     = "🔵" if n["type"] == "Individual" else ("🟢" if n["type"] == "Pooled" else "📂")
            is_direct = n["type"] == "Direct"
            mol_markers = n.get("mol_markers", [])
            mol_code    = n.get("mol_code", "000")
            is_editing  = st.session_state["editing_idx"] == i

            with st.container():
                if is_direct:
                    st.markdown(f"""
                    <div class="assoc-row">
                        <div style="display:flex;align-items:center;gap:6px;margin-bottom:6px;flex-wrap:wrap;">
                            <span style="font-size:0.75rem;font-weight:700;color:#888888;">#{i+1}</span>
                            <span style="font-size:0.75rem;font-weight:700;color:#888888;text-transform:uppercase;letter-spacing:1px;">{badge} {n['type']}</span>
                        </div>
                        <div style="font-family:'JetBrains Mono',monospace;font-size:0.85rem;color:#2c6fad;margin-bottom:3px;">{n.get("orig_r1", n["r1"])}</div>
                        <div style="font-family:'JetBrains Mono',monospace;font-size:0.85rem;color:#6f42c1;">{n.get("orig_r2", n["r2"])}</div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    mol_badge_html = f'<span style="font-size:0.75rem;color:#7c3aed;background:#faf5ff;border:1px solid #d8b4fe;border-radius:3px;padding:2px 8px;margin-left:4px;">🧬 mol:{mol_code} ({", ".join(mol_markers) if mol_markers else "aucun"})</span>'
                    st.markdown(f"""
                    <div class="assoc-row">
                        <div style="display:flex;align-items:center;gap:6px;margin-bottom:6px;flex-wrap:wrap;">
                            <span style="font-size:0.75rem;font-weight:700;color:#888888;">#{i+1}</span>
                            <span style="font-size:0.75rem;font-weight:700;color:#888888;text-transform:uppercase;letter-spacing:1px;">{badge} {n['type']}</span>
                            {mol_badge_html}
                        </div>
                        <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:3px;">
                            <span style="font-family:'JetBrains Mono',monospace;font-size:0.82rem;color:#888888;">{n.get("orig_r1", n["r1"])}</span>
                            <span style="color:#2c6fad;font-size:1rem;">→</span>
                            <span style="font-family:'JetBrains Mono',monospace;font-size:0.85rem;color:#2c6fad;font-weight:600;">{n["r1"]}</span>
                        </div>
                        <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;">
                            <span style="font-family:'JetBrains Mono',monospace;font-size:0.82rem;color:#888888;">{n.get("orig_r2", n["r2"])}</span>
                            <span style="color:#6f42c1;font-size:1rem;">→</span>
                            <span style="font-family:'JetBrains Mono',monospace;font-size:0.85rem;color:#6f42c1;font-weight:600;">{n["r2"]}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    if mol_markers is not None:
                        render_mol_bitcode(mol_markers, mol_code)

                btn_col1, btn_col2 = st.columns([1, 8])
                with btn_col1:
                    edit_label = "✏️ Modifier" if not is_editing else "✕ Fermer"
                    if st.button(edit_label, key=f"edit_btn_{i}"):
                        st.session_state["editing_idx"] = None if is_editing else i
                        st.rerun()

                if is_editing:
                    st.markdown(f"""
                    <div style="background:#f0f6ff;border:1px solid #bdd5ef;border-radius:4px;padding:14px 16px;margin-bottom:8px;">
                        <div style="font-size:0.75rem;font-weight:700;color:#2c6fad;letter-spacing:1px;text-transform:uppercase;margin-bottom:10px;">
                            ✏️ Édition — Échantillon #{i+1}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    ec1, ec2, ec3, ec4 = st.columns(4)
                    cur_year    = n["base"][0:2]   if len(n["base"]) >= 2  else ""
                    cur_country = n["base"][2:4]   if len(n["base"]) >= 4  else ""
                    cur_state   = n["base"][4:6]   if len(n["base"]) >= 6  else ""
                    cur_day     = n["base"][6:8]   if len(n["base"]) >= 8  else ""
                    cur_treat   = n["base"][8:9]   if len(n["base"]) >= 9  else ""

                    e_year    = ec1.text_input("Année",         max_chars=2, value=cur_year,    key=f"e_year_{i}")
                    e_country = ec2.text_input("Pays",          max_chars=2, value=cur_country, key=f"e_country_{i}")
                    e_state   = ec3.text_input("État/Province", max_chars=2, value=cur_state,   key=f"e_state_{i}")
                    e_day     = ec4.text_input("Jour",          max_chars=2, value=cur_day,     key=f"e_day_{i}")

                    saved_drug  = n.get("drug", DRUG_OPTIONS[0])
                    drug_index  = DRUG_OPTIONS.index(saved_drug) if saved_drug in DRUG_OPTIONS else 0
                    e_drug      = st.selectbox("💊 Médicament", DRUG_OPTIONS, index=drug_index, key=f"e_drug_{i}")
                    if e_drug != DRUG_OPTIONS[0]:
                        e_treat = DRUG_CODES[e_drug]
                        st.markdown(f'<div style="background:#f0fdf4;border:1px solid #a8d5b5;border-radius:3px;padding:5px 12px;margin-bottom:8px;font-size:0.85rem;color:#155724;">✅ Code traitement : <strong>{e_treat}</strong></div>', unsafe_allow_html=True)
                    else:
                        e_treat = st.text_input("Code traitement (1 lettre)", max_chars=1, value=cur_treat, key=f"e_treat_{i}")

                    st.markdown('<div style="font-size:0.85rem;font-weight:600;color:#555555;margin-top:6px;margin-bottom:4px;">🧬 Marqueurs moléculaires</div>', unsafe_allow_html=True)
                    e_mol_sel = st.multiselect(
                        "Marqueurs", options=MOL_MARKERS_LIST,
                        default=n.get("mol_markers", []),
                        key=f"e_mol_{i}", label_visibility="collapsed"
                    )
                    e_mol_code = compute_mol_code(e_mol_sel)
                    render_mol_bitcode(e_mol_sel, e_mol_code)

                    cur_type = n.get("type", "Individual")
                    b = n["base"]
                    if cur_type == "Individual":
                        cur_sid_ind = b[9:13]  if len(b) >= 13 else ""
                        cur_proc    = b[19:20] if len(b) >= 20 else (b[-1] if b else "1")
                        cur_sid3 = ""; cur_npool = ""
                    elif cur_type == "Pooled":
                        cur_sid3    = b[9:12]  if len(b) >= 12 else ""
                        cur_npool   = b[13:15] if len(b) >= 15 else ""
                        cur_proc    = b[19:20] if len(b) >= 20 else (b[-1] if b else "1")
                        cur_sid_ind = ""
                    else:
                        cur_sid_ind = ""; cur_sid3 = ""; cur_npool = ""; cur_proc = "1"
                    if cur_proc == "Z": cur_proc = "1"
                    elif cur_proc == "Y": cur_proc = "2"
                    elif cur_proc not in ("1", "2"): cur_proc = "1"

                    e_type = st.radio("Type", ["Individual", "Pooled"], horizontal=True,
                                      key=f"e_type_{i}",
                                      index=0 if cur_type in ("Individual","Direct") else 1)
                    if e_type == "Individual":
                        e_sid  = st.text_input("Sample ID Individual (4 chiffres)", max_chars=4,
                                               value=cur_sid_ind, key=f"e_sid_{i}")
                        e_proc = st.radio("Processé", ["1", "2"], horizontal=True, key=f"e_proc_{i}",
                                          index=0 if cur_proc != "2" else 1)
                    else:
                        ci1, ci2, ci3 = st.columns(3)
                        e_sid3  = ci1.text_input("ID Pooled (3 chiffres)", max_chars=3, value=cur_sid3, key=f"e_sid3_{i}")
                        ci2.text_input("P", disabled=True, key=f"e_pool_p_{i}")
                        e_npool = ci3.text_input("Nb dans pool (2 chiffres)", max_chars=2, value=cur_npool, key=f"e_npool_{i}")
                        e_proc  = st.radio("Processé", ["1", "2"], horizontal=True, key=f"e_proc_{i}",
                                           index=0 if cur_proc != "2" else 1)

                    if st.button("💾 Sauvegarder les modifications", key=f"save_edit_{i}"):
                        err = validate_common(e_year, e_country, e_state, e_day, e_treat, e_mol_code, e_proc)
                        new_base = ""
                        if not err:
                            if e_type == "Individual":
                                e_sid_val = locals().get("e_sid", "")
                                if only_digits(e_sid_val, 4):
                                    new_base = build_base(e_year, e_country, e_state, e_day, e_treat, e_sid_val, "Pf" + "B" + e_mol_code + e_proc)
                                else:
                                    err.append("Sample ID Individual : exactement 4 chiffres")
                            else:
                                e_sid3_val  = locals().get("e_sid3", "")
                                e_npool_val = locals().get("e_npool", "")
                                if only_digits(e_sid3_val, 3) and only_digits(e_npool_val, 2):
                                    new_base = build_base(e_year, e_country, e_state, e_day, e_treat, e_sid3_val + "P" + e_npool_val, "B" + e_mol_code + e_proc)
                                else:
                                    err.append("Identifiants Pooled invalides")
                        if err:
                            for e in err: st.error(e)
                        elif new_base:
                            new_r1 = build_new_filename(new_base, n.get("orig_r1", n["r1"]))
                            new_r2 = build_new_filename(new_base, n.get("orig_r2", n["r2"]))
                            names[i].update({
                                "base": new_base, "type": e_type,
                                "r1": new_r1, "r2": new_r2,
                                "mol_markers": e_mol_sel, "mol_code": e_mol_code, "drug": e_drug,
                            })
                            st.session_state["sample_names"] = names
                            st.session_state["editing_idx"]  = None
                            st.toast(f"Échantillon #{i+1} mis à jour ✓", icon="✅")
                            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        col_launch, col_reset = st.columns([3, 1])
        with col_launch:
            if st.button("▶  Lancer l'analyse bioinformatique", width='stretch'):
                st.session_state["trigger_run"] = True
        with col_reset:
            if st.button("🔄 Tout recommencer", width='stretch'):
                reset_state()
                st.rerun()

        if st.session_state.get("trigger_run", False):
            st.session_state["trigger_run"]    = False
            st.session_state["run_id"]         = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.session_state["running"]        = True
            st.session_state["pipeline_done"]  = False
            st.session_state["zip_created"]    = False
            st.session_state["show_download"]  = False
            st.session_state["docker_started"] = False

            run_id = st.session_state["run_id"]
            paths  = get_run_paths(run_id)
            paths["input"].mkdir(parents=True, exist_ok=True)
            paths["output"].mkdir(parents=True, exist_ok=True)
            paths["logs"].mkdir(parents=True, exist_ok=True)

            CHUNK = 64 * 1024 * 1024  # 64 Mo par chunk — évite le pic mémoire
            for n in names:
                file_r1 = n.get("file_r1")
                file_r2 = n.get("file_r2")
                if file_r1:
                    file_r1.seek(0)
                    with open(paths["input"] / safe_filename(n["r1"]), "wb") as f:
                        while chunk := file_r1.read(CHUNK):
                            f.write(chunk)
                if file_r2:
                    file_r2.seek(0)
                    with open(paths["input"] / safe_filename(n["r2"]), "wb") as f:
                        while chunk := file_r2.read(CHUNK):
                            f.write(chunk)

            log_file = paths["logs"] / "pipeline.log"
            with open(log_file, "w") as lf:
                subprocess.Popen(
                    [
                        "docker", "run", "--rm",
                        "--label", f"mars_run_id={run_id}", 
                        "-v", f"{paths['base']}:/pipeline",
                        "-v", f"{REPO_DIR}/pf_3D7_Ref:/ref:ro",
                        "-v", f"{REPO_DIR}/pf_3D7_snpEff_db:/snpeff_db:ro",
                        "-v", f"{REPO_DIR}:/app",
                        "-v", f"{REPO_DIR}/pipeline_python.py:/app/pipeline_python.py",
                        "bioinfo_pipeline"
                    ],
                    stdout=lf, stderr=lf
                )

            st.session_state["log_file"]          = str(log_file)
            st.session_state["launch_time"]       = datetime.now().timestamp()
            st.session_state["names_for_history"] = names
            save_run_to_history(run_id, names, "running", None, None)
            st.rerun()

    # ════════════════════════════════════════════
    # ÉTAPE 3 — PIPELINE EN COURS
    # ════════════════════════════════════════════
    if st.session_state.get("running"):
             render_progress_tracker(
                 get_run_paths_fn  = get_run_paths,
                 save_history_fn   = save_run_to_history,
                 notify_fn         = notify_pipeline_done,
             )

    # ════════════════════════════════════════════
    # ARRÊT / ÉCHEC
    # ════════════════════════════════════════════
    if (not st.session_state["running"]
            and not st.session_state["pipeline_done"]
            and st.session_state["nomenclature_done"]
            and st.session_state["run_id"] is not None):
        st.markdown("---")
        st.error("🛑 L'analyse a été arrêtée ou a échoué.")
        col_retry, col_home = st.columns(2)
        with col_retry:
            if st.button("🔄 Relancer l'analyse", width='stretch'):
                st.session_state["run_id"]        = None
                st.session_state["zip_created"]   = False
                st.session_state["show_download"] = False
                st.session_state["log_file"]      = ""
                st.rerun()
        with col_home:
            if st.button("🏠 Retour à l'accueil", key="home_after_stop", width='stretch'):
                reset_state()
                st.rerun()

    # ════════════════════════════════════════════
    # ÉTAPE 4 — RÉSULTATS
    # ════════════════════════════════════════════
    if st.session_state["pipeline_done"] and st.session_state["run_id"]:
        names  = st.session_state.get("sample_names", [])
        paths  = get_run_paths(st.session_state["run_id"])
        run_id = st.session_state["run_id"]

        st.markdown("---")
        st.markdown('<div class="section-label">Étape 4 — Résultats</div>', unsafe_allow_html=True)

        history   = load_history()
        run_entry = next((h for h in history if h["run_id"] == run_id), None)
        duration_sec = run_entry.get("duration_sec") if run_entry else None
        n_samples    = run_entry.get("samples", len(names)) if run_entry else len(names)

        if duration_sec is not None:
            dur_min = duration_sec // 60; dur_sec = duration_sec % 60
            dur_str = f"{dur_min} min {dur_sec} s" if dur_min > 0 else f"{dur_sec} s"
        else:
            lt = st.session_state.get("launch_time", 0)
            e  = int(datetime.now().timestamp() - lt) if lt else 0
            dur_str = f"{e//60} min {e%60} s" if e else "—"

        st.markdown(f"""
        <div style="background:#edfaf0;border:1px solid #a8d5b5;border-radius:6px;padding:28px 32px;margin-bottom:20px;display:flex;align-items:center;gap:20px;">
            <div style="font-size:2.5rem;">✅</div>
            <div>
                <div style="font-family:'Source Sans Pro',sans-serif;font-size:1.4rem;color:#1a2a4a;margin-bottom:4px;font-weight:700;">Analyse terminée avec succès</div>
                <div style="font-size:0.85rem;color:#155724;margin-bottom:6px;">Run {run_id} · {n_samples} échantillon(s) · ⏱ {dur_str}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        zip_path = paths["zip"]
        if not zip_path.exists() and paths["output"].exists():
            with zipfile.ZipFile(zip_path, "w") as zipf:
                for root, _, files in os.walk(paths["output"]):
                    for file in files:
                        full_path = os.path.join(root, file)
                        zipf.write(full_path, os.path.relpath(full_path, paths["output"]))
        if zip_path.exists():
            with open(zip_path, "rb") as f:
                st.download_button("💾 Télécharger les résultats (ZIP)", data=f, file_name=f"resultats_{run_id}.zip", mime="application/zip", width='stretch')
        else:
            st.warning("⚠️ Fichier ZIP introuvable.")

        qc_data = get_qc_data_for_run(run_id)
        if qc_data:
            st.markdown("---")
            st.markdown('<div class="section-label">Aperçu QC FastQC</div>', unsafe_allow_html=True)
            sel_qc = st.selectbox("Sélectionner un échantillon", list(qc_data.keys()), key="qc_inline_sel")
            if sel_qc:
                d = qc_data[sel_qc]; m = d.get("txt_metrics", {})
                if m:
                    c1, c2, c3 = st.columns(3)
                    for col, label, value in [(c1,"Lectures totales",m.get("total_reads","—")),(c2,"Longueur lecture",m.get("read_length","—")),(c3,"% GC",m.get("gc_pct","—"))]:
                        col.markdown(f'<div class="qc-metric"><div class="qc-label">{label}</div><div class="qc-value">{value}</div></div>', unsafe_allow_html=True)
                    mods = m.get("modules", {})
                    key_modules = ["Basic Statistics","Per base sequence quality","Per sequence quality scores","Sequence Duplication Levels"]
                    mod_html = " &nbsp; ".join(f'{get_module_badge(mods[mod])} <span style="font-size:0.75rem;color:#495057;">{mod}</span>' for mod in key_modules if mod in mods)
                    if mod_html:
                        st.markdown(f'<div style="margin-top:12px;display:flex;flex-wrap:wrap;gap:8px;align-items:center;">{mod_html}</div>', unsafe_allow_html=True)
            if st.button("📊 Voir le dashboard QC complet"):
                st.session_state["prev_page"]   = "pipeline"
                st.session_state["qc_run_id"]   = run_id
                st.session_state["active_page"] = "qc_detail"
                st.rerun()

        # Bouton Résultats moléculaires
        if st.button("🧬 Voir les résultats moléculaires", width='stretch'):
            st.session_state["prev_page"]      = st.session_state.get("active_page", "home")
            st.session_state["results_run_id"] = run_id
            st.session_state["active_page"]    = "results"
            st.rerun()

        st.markdown("---")
        if st.button("🏠 Nouvelle analyse", width='stretch'):
            reset_state()
            st.rerun()
